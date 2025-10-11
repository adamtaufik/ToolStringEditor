import os
import time
import win32com.client
from PyQt6.QtGui import QPixmap
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Border, Side, Alignment
from PIL import Image as PILImage

from utils.loading_worker import LoadingWorker
from features.editors.logic_image_processing import combine_tool_images, expand_and_center_images, remove_white_background
from utils.logic_utils import get_number
from ui.components.toolstring_editor.tool_widget import ToolWidget
from ui.windows.ui_messagebox_window import MessageBoxWindow
from utils.check_file import is_file_open
from utils.path_finder import get_icon_path, get_tool_image_path
from io import BytesIO


def export_configuration(main_window):
    """Exports the current tool string to an Excel file in a separate thread."""
    from ui.components.toolstring_editor.workers import ExportWorker

    # Suggest filename based on client + location or last saved file
    default_name = main_window.current_file_name or f"{main_window.location.text()}_{main_window.well_no.text()}_{main_window.operation_details.text()}".replace(
        " ", "_")
    default_path = os.path.join(os.getcwd(), default_name.replace(".json", ""))  # Set default path

    # ✅ Check if DropZone is empty
    if not main_window.drop_zone.tool_widgets:
        MessageBoxWindow.message_simple(main_window,
                                        "Export Error",
                                        "The tool string is empty. Please add tools before exporting.",
                                        QMessageBox.Icon.Warning)

        return  # ✅ Stop export process if empty

    file_dialog = QFileDialog()
    excel_path, _ = file_dialog.getSaveFileName(main_window, "Export Tool String", default_path, "Excel Files (*.xlsx)")

    if not excel_path:
        return  # ✅ Exit if user cancels

    # Before saving
    if is_file_open(excel_path):
        # print(f"⚠️ The file {excel_path} is open in Excel. Please close it and try again.")

        print(f"Excel is currently open. Please close any Excel windows and try again.")
        MessageBoxWindow.message_simple(main_window,
                                        "Export Error",
                                        f"Excel is currently open. Please close any Excel windows and try again.",
                                        QMessageBox.Icon.Warning)

        return  # Stop execution

    pdf_path = excel_path.replace(".xlsx", ".pdf")
    final_directory = os.path.dirname(excel_path)  # ✅ Extract directory

    # ✅ **Start Loading Animation in a Separate Thread**
    main_window.loading_worker = LoadingWorker(main_window)
    main_window.loading_worker.start()

    # ✅ **Start Export in a Background Thread**

    main_window.export_thread = ExportWorker(main_window, excel_path, pdf_path, final_directory)
    main_window.export_thread.finished.connect(lambda: on_export_finished(main_window, final_directory))
    main_window.export_thread.start()


def on_export_finished(main_window, final_directory):
    """Called when the export thread is finished."""

    main_window.loading_worker.stop_dialog()

    response = MessageBoxWindow.message_yes_no(main_window,
                                               "Export Successful",
                                               f"Tool string exported successfully!\n\n📂 Folder location:\n{final_directory}\n\nWould you like to open the folder?",
                                               QMessageBox.Icon.Information)

    # ✅ Open actual save directory if user clicks "Yes"
    if response == QMessageBox.StandardButton.Yes:
        try:
            os.startfile(final_directory)  # ✅ Opens the correct folder
        except Exception as e:
            print(f"⚠️ ERROR: Unable to open folder: {e}")

def export_to_excel(excel_path, pdf_path, client_name, location, well_no, max_angle, well_type, date, operation_details, comments, drop_zone):
    """Exports tool string configuration to Excel and PDF."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Tool String"
    last_column = 'H'
    last_row = 52

    if drop_zone.layout.count() > 18:
        last_row += 2 * (drop_zone.layout.count() - 18) - 1

    png_path = excel_path.replace(".xlsx", ".png")

    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # --- Title ---
    ws["A1"] = "TOOL STRING SCHEMATIC"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A4"] = "Operation Details"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = operation_details

    # --- Client Info ---
    client_info = [
        ["Client Name", "", "Location", "", "Well No.", "Well Type", "Max Angle", "Date"],
        [client_name, "", location, "", well_no, well_type, max_angle, date]
    ]
    for row_idx, row_data in enumerate(client_info, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx == 2:
                cell.font = Font(bold=True)

    data, tool_images = extract_tool_data(drop_zone)
    print("Tool data extracted successfully!")

    headers = ["Diagram", "", "Description", "OD (in)", "Top Connection", "Bottom Connection", "Length (ft)", "Weight (lbs)"]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_length = 0.0
    total_weight = 0.0
    od_list = []

    for row_data in data:
        if len(data) < 19:
            ws.append([])
        ws.append(row_data)
        od_list.append(row_data[3])
        total_length += get_number(row_data[6])
        total_weight += row_data[7]

    max_od = max(od_list)
    cell_remarks_title = 'C' + str(last_row - 5)
    cell_remarks_content = 'C' + str(last_row - 4)
    ws[cell_remarks_title] = "Remarks"
    ws[cell_remarks_title].font = Font(bold=True)

    # --- Apply Borders & Alignment ---
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=8):
        for cell in row:
            row_idx = cell.row
            col_letter = cell.column_letter

            # --- Apply only vertical borders inside C7:H44 ---
            if ("C" <= col_letter <= "H") and (7 <= row_idx <= 44):
                # Determine if cell is on the outer border of the region
                is_top = (row_idx == 7)
                is_bottom = (row_idx == 44)
                is_left = (col_letter == "C")
                is_right = (col_letter == "H")

                # Only vertical borders inside, full borders on edges
                cell.border = Border(
                    left=thin_side if is_left or True else None,
                    right=thin_side if is_right or True else None,
                    top=thin_side if is_top else None,
                    bottom=thin_side if is_bottom else None
                )
            else:
                # Default full border for other areas
                cell.border = thin_border

            cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Auto column width ---
    for col in ws.iter_cols(min_col=3, max_col=8):
        max_length = 0
        try:
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        except:
            pass

    # --- Summary Rows ---
    for column in ['C', 'D', 'E', 'F', 'G', 'H']:
        cell1 = column + str(last_row - 7)
        cell2 = column + str(last_row - 6)
        ws[cell1].font = Font(bold=True)
        ws[cell2].font = Font(bold=True)
        if column == 'C':
            ws[cell1] = "Max OD (in)"
            ws[cell2] = "Max OD (mm)"
        elif column == 'D':
            ws[cell1] = max_od
            ws[cell2] = round(max_od * 25.4, 1)
        if column == 'E':
            ws[cell1] = "Total Length (ft) & Weight (lbs)"
            ws[cell2] = "Total Length (m) & Weight (kg)"
            ws.merge_cells(f"{cell1}:F{last_row-7}")
            ws.merge_cells(f"{cell2}:F{last_row-6}")
        elif column == 'G':
            ws[cell1] = total_length
            ws[cell2] = round(total_length * 0.3048, 1)
        elif column == 'H':
            ws[cell1] = total_weight
            ws[cell2] = round(total_weight * 0.453592, 1)

    # --- Merging ---
    ws.merge_cells(f"A1:{last_column}1")
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")
    ws.merge_cells("A6:B6")
    ws.merge_cells("C2:D2")
    ws.merge_cells("C3:D3")
    ws.merge_cells(f"A4:{last_column}4")
    ws.merge_cells(f"A5:{last_column}5")
    ws.merge_cells(f"A7:B{last_row}")
    ws.merge_cells(f"{cell_remarks_title}:{last_column}{last_row - 5}")
    ws.merge_cells(f"{cell_remarks_content}:{last_column}{last_row}")

    ws[cell_remarks_content] = comments.replace("\n", "\n")
    ws[cell_remarks_content].alignment = Alignment(wrapText=True, horizontal="left")

    # --- Insert Image ---
    cell = "B8"
    if tool_images:
        centered_images = expand_and_center_images(tool_images, 80, True)
        tool_image = combine_tool_images(centered_images)
        tool_image = remove_white_background(tool_image)
        tool_image.save(png_path)
        pil_img = PILImage.open(png_path)
        max_height = 850
        if pil_img.height > max_height:
            scale_factor = max_height / pil_img.height
            new_width = int(pil_img.width * scale_factor)
            pil_img = pil_img.resize((new_width, max_height), PILImage.LANCZOS)
            pil_img.save(png_path)
        img = ExcelImage(png_path)
        ws.add_image(img, cell)

    ws.column_dimensions['A'].width = 4.51
    ws.column_dimensions['B'].width = 4.51 + img.width / 7

    ws.row_dimensions[1].height = 55
    row_1_height = ws.row_dimensions[1].height
    logo_path = get_icon_path('logo_report')
    logo_img = PILImage.open(logo_path)
    max_logo_height = row_1_height * 1.25
    scale_factor = min(1, max_logo_height / logo_img.height)
    new_width = int(logo_img.width * scale_factor)
    new_height = int(logo_img.height * scale_factor)
    logo_img = logo_img.resize((new_width, new_height), PILImage.LANCZOS)
    img_bytes = BytesIO()
    logo_img.save(img_bytes, format="PNG")
    img_bytes.seek(0)
    excel_logo = ExcelImage(img_bytes)
    ws.add_image(excel_logo, "A1")

    footer_cell = last_column + str(last_row + 2)
    ws[footer_cell] = "This report was computer generated using Deleum Tool String Editor"
    ws[footer_cell].alignment = Alignment(horizontal="right",)

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{last_column}{ws.max_row}"

    wb.save(excel_path)
    print(f"✅ Excel export successful: {excel_path}")


def export_to_pdf(excel_path, pdf_path):
    """Converts an Excel file to PDF and ensures Excel closes properly."""
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # ✅ **Use absolute paths to prevent directory issues**
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Sheets(1)

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        ws.PageSetup.PaperSize = 9
        ws.PageSetup.Orientation = 1

        # ✅ **Ensure the file is closed before exporting to PDF**
        time.sleep(2)
        
        wb.ExportAsFixedFormat(0, pdf_path)
        print(f"✅ Successfully exported PDF: {pdf_path}")

    except Exception as e:
        print(f"❌ Failed to export PDF: {e}")
    finally:
        try:
            wb.Close(SaveChanges=False)
            excel.Quit()
            print("✅ Excel closed successfully.")
        except:
            print("⚠️ Excel instance may still be running.")

    print(f"Excel Path: {excel_path}")
    print(f"PDF Path: {pdf_path}")

def extract_tool_data(drop_zone):
    """Extracts tool details and images from the drop zone."""
    data = []
    tool_images = []

    for i in range(drop_zone.layout.count()):
        widget = drop_zone.layout.itemAt(i).widget()

        # ✅ Ensure the widget is a ToolWidget
        if not isinstance(widget, ToolWidget):
            continue

        # **Extract Tool Information**
        tool_name = widget.tool_name
        size = widget.nominal_size_selector.currentText()
        od = get_number(widget.od_label.text())  # Ensure OD is numeric
        top_connection = widget.top_connection_label.text()
        lower_connection = widget.lower_connection_label.currentText()
        length = widget.length_label.text()
        weight = get_number(widget.weight_label.text())  # Ensure weight is numeric

        # **Append tool details to the list**
        data.append(["", "", f"{tool_name} ({size})", od, top_connection, lower_connection, length, weight])

        # **Retrieve Tool Image**

        image_path = get_tool_image_path(tool_name)
        pixmap = QPixmap(image_path)

        # **Store tool image only if it's valid**
        if widget.image_label.pixmap() and not widget.image_label.pixmap().isNull():
            tool_images.append(pixmap.toImage())

    return data, tool_images

