import os
import time
import win32com.client
from PyQt6.QtGui import QPixmap
from PyQt6.QtWidgets import QMessageBox, QFileDialog
from openpyxl import Workbook

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Border, Side, Alignment
from PIL import Image as PILImage
from openpyxl.utils import range_boundaries, get_column_letter

from features.pce_editor.loading_worker import LoadingWorker
from features.pce_editor.logic_image_processing import combine_tool_images, expand_and_center_images, remove_white_background
from features.pce_editor.logic_utils import get_number
from ui.components.pce_editor.tool_widget import ToolWidget
from ui.windows.ui_messagebox_window import MessageBoxWindow
from utils.check_file import is_file_open
from utils.path_finder import get_icon_path, get_pce_image_path
from io import BytesIO


def export_configuration(main_window):
    """Exports the current tool string to an Excel file in a separate thread."""
    from ui.components.pce_editor.workers import ExportWorker

    # Suggest filename based on client + location or last saved file
    default_name = main_window.current_file_name or f"{main_window.location.text()}_{main_window.well_no.text()}_{main_window.operation_details.text()}".replace(
        " ", "_")
    default_path = os.path.join(os.getcwd(), default_name.replace(".json", ""))  # Set default path

    # ✅ Check if DropZone is empty
    if not main_window.drop_zone.tool_widgets:
        MessageBoxWindow.message_simple(main_window,
                                        "Export Error",
                                        "The tool string is empty. Please add tools before exporting.",
                                        QMessageBox.Icon.Warning)

        return  # ✅ Stop export process if empty

    file_dialog = QFileDialog()
    excel_path, _ = file_dialog.getSaveFileName(main_window, "Export Tool String", default_path, "Excel Files (*.xlsx)")

    if not excel_path:
        return  # ✅ Exit if user cancels

    # Before saving
    if is_file_open(excel_path):
        # print(f"⚠️ The file {excel_path} is open in Excel. Please close it and try again.")

        print(f"Excel is currently open. Please close any Excel windows and try again.")
        MessageBoxWindow.message_simple(main_window,
                                        "Export Error",
                                        f"Excel is currently open. Please close any Excel windows and try again.",
                                        QMessageBox.Icon.Warning)

        return  # Stop execution

    pdf_path = excel_path.replace(".xlsx", ".pdf")
    final_directory = os.path.dirname(excel_path)  # ✅ Extract directory

    # ✅ **Start Loading Animation in a Separate Thread**
    main_window.loading_worker = LoadingWorker(main_window)
    main_window.loading_worker.start()

    # ✅ **Start Export in a Background Thread**

    main_window.export_thread = ExportWorker(main_window, excel_path, pdf_path, final_directory)
    main_window.export_thread.finished.connect(lambda: on_export_finished(main_window, final_directory))
    main_window.export_thread.start()


def on_export_finished(main_window, final_directory):
    """Called when the export thread is finished."""

    main_window.loading_worker.stop_dialog()

    response = MessageBoxWindow.message_yes_no(main_window,
                                               "Export Successful",
                                               f"Tool string exported successfully!\n\n📂 Folder location:\n{final_directory}\n\nWould you like to open the folder?",
                                               QMessageBox.Icon.Information)

    # ✅ Open actual save directory if user clicks "Yes"
    if response == QMessageBox.StandardButton.Yes:
        try:
            os.startfile(final_directory)  # ✅ Opens the correct folder
        except Exception as e:
            print(f"⚠️ ERROR: Unable to open folder: {e}")

def export_to_excel(excel_path, pdf_path, client_name, location, well_no, date, operation_details, comments, drop_zone):
    """Exports tool string configuration to Excel and PDF."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Tool String"
    last_column = 'I'
    last_row = 52

    if drop_zone.layout.count() > 18:
        last_row += 2 * (drop_zone.layout.count() - 18) - 1

    png_path = excel_path.replace(".xlsx",".png")

    # **Define border style**
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # **Set title**
    ws["A1"] = "PCE STACK-UP"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A4"] = "Operation Details"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = operation_details

    ws["A4"].alignment = Alignment(horizontal="center")
    ws["A5"].alignment = Alignment(horizontal="center")

    # **Client Information Section**
    client_info = [
        ["Client Name", "", "Location", "", "", "Well No.", "", "Date"],
        [client_name, "", location, "", "", well_no, "", date]
    ]

    for row_idx, row_data in enumerate(client_info, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx == 2:
                cell.font = Font(bold=True)

    print('trying to extract')
    try:
        # **Extract Tool Data**
        data, tool_images = extract_tool_data(drop_zone)  # ✅ Call helper function

        tool_images.append(QPixmap(get_pce_image_path("Christmas Tree")).toImage())

        # how tall each source image was padded to before stacking (your expand_and_center_images call)
        SOURCE_SEGMENT_PX = 150  # must match your expand_and_center_images(..., 150, ...)

        tool_count = len(data)  # rows we will place
        segment_count = len(tool_images)  # includes the extra "Christmas Tree"

        DATA_START_ROW = 7  # first table data row (headers at 6)
        ROW_PX = 24  # target row height in *pixels* (~18pt)
        ROW_PT = ROW_PX * 0.75  # Excel uses points; 1 pt ≈ 1.333 px

        # ensure all data rows use the same height (so conversion pixels->rows is stable)
        # we’ll fill the exact range later once we know last_row
        rows_with_custom_height = []
    except Exception as e:
        print(e)

    # ✅ You can now use `tool_images` for image processing
    print("Tool data extracted successfully!")

    try:
        # **Table Headers**
        headers = ["Diagram", "", "Description", "ID (in)", "Connections", "Length (ft)", "Service", "WP (psi)", "Weight (kg)"]

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        # === Calculate Totals (just compute numbers here) ===
        total_length = 0.0
        total_weight = 0.0
        for row_data in data:
            total_length += get_number(row_data[5])
            total_weight += row_data[8]

    except Exception as e:
        print(e)

    # === Insert Tool String Image (dpi-aware) and place data rows aligned to centers ===
    cell = "B7"  # align the image top with the first data row band
    have_image = False
    img = None  # so we can test later

    if tool_images:
        centered_images = expand_and_center_images(tool_images, SOURCE_SEGMENT_PX, True)
        tool_image = combine_tool_images(centered_images)
        tool_image = remove_white_background(tool_image)

        tool_image.save(png_path)

        pil_img = PILImage.open(png_path)
        max_height = 1000
        if pil_img.height > max_height:
            scale_factor = max_height / pil_img.height
            pil_img = pil_img.resize(
                (int(pil_img.width * scale_factor), int(pil_img.height * scale_factor)),
                PILImage.LANCZOS
            )
            pil_img.save(png_path)

        img = ExcelImage(png_path)
        ws.add_image(img, cell)
        have_image = True

        # widen columns for the diagram
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 3 + img.width / 7

        # --- compute per-tool center rows ---
        final_total_height_px = pil_img.height
        final_segment_px = final_total_height_px / float(segment_count)  # all segments same height
        rows_per_segment = final_segment_px / ROW_PX
        center_offset_rows = rows_per_segment / 2.0

        center_rows = []
        last_used = DATA_START_ROW - 1
        for i in range(tool_count):  # ignore final tree segment for data rows
            r = int(round(DATA_START_ROW + center_offset_rows + i * rows_per_segment))
            if r <= last_used:
                r = last_used + 1
            center_rows.append(r)
            last_used = r

        # write each data row at its computed center row
        first_data_row = None
        last_data_row = None
        for idx, row_data in enumerate(data):
            r = center_rows[idx]
            first_data_row = r if first_data_row is None else min(first_data_row, r)
            last_data_row = r if last_data_row is None else max(last_data_row, r)

            ws.row_dimensions[r].height = ROW_PT
            rows_with_custom_height.append(r)
            for col_num, value in enumerate(row_data, start=1):
                # Enable line breaks in the connection column
                if col_num == 5:  # Column E (connections)
                    top = value.split("\n")[0]
                    bottom = value.split("\n")[1]
                    if top == '-':
                        cell = ws.cell(row=r, column=col_num, value=bottom)
                    else:
                        cell = ws.cell(row=r, column=col_num, value=top)
                        cell = ws.cell(row=r+1, column=col_num, value=bottom)
                else:
                    cell = ws.cell(row=r, column=col_num, value=value)



    else:
        # fallback if no image
        first_data_row = DATA_START_ROW
        last_data_row = DATA_START_ROW + len(data) - 1

    # === Decide final last_row now that we know where data ends ===
    summary_pad = 10
    last_row = max(last_row, last_data_row + summary_pad)

    # normalize row heights across the whole data band for even borders
    for r in range(DATA_START_ROW, last_row + 1):
        if r not in rows_with_custom_height:
            ws.row_dimensions[r].height = ROW_PT

    # If you want to ensure the diagram column width even without an image
    if not have_image:
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18

    # === Remarks (after we know last_row) ===
    cell_remarks_title = f"C{last_row - 3}"
    cell_remarks_content = f"C{last_row - 2}"
    ws[cell_remarks_title] = "Remarks"
    ws[cell_remarks_title].font = Font(bold=True)
    ws[cell_remarks_content] = comments.replace("\n", "\n")
    ws[cell_remarks_content].alignment = Alignment(wrapText=True, horizontal="left")

    print(275)
    # === Merges (use FINAL last_row) ===
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")
    ws.merge_cells("A6:B6")  # (removed duplicate)
    ws.merge_cells(f"A7:B{last_row}")  # diagram column spans entire data+summary area

    # === Auto width for C..H AFTER values are in ===
    for col in ws.iter_cols(min_col=3, max_col=10):
        max_length = 0
        try:
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        except:
            pass

    thin = Side(style="thin")

    print(296)
    # === Table headers (already written at row 6); now add summary cells using FINAL last_row ===
    for column in ['C', 'D', 'E', 'F', 'G', 'H', 'J']:
        cell1 = f"{column}{last_row - 7}"
        cell2 = f"{column}{last_row - 6}"
        cell3 = f"{column}{last_row - 5}"
        cell4 = f"{column}{last_row - 4}"
        ws[cell1].font = Font(bold=True)
        ws[cell2].font = Font(bold=True)
        ws[cell3].font = Font(bold=True)
        ws[cell4].font = Font(bold=True)
        ws[cell1].alignment = Alignment(vertical="center")
        ws[cell2].alignment = Alignment(vertical="center")
        ws[cell3].alignment = Alignment(vertical="center")
        ws[cell4].alignment = Alignment(vertical="center")
        if column == 'C':
            ws[cell1] = "Minimum ID (in)"
            ws[cell3] = "Maximum ID (in)"
        if column == 'E':
            ws[cell1] = "Bott. of Stuffing Box to Top of BOP (ft)"
            ws[cell2] = "Bott. of BOP to Top of Tree (ft)"
            ws[cell3] = "Maximum tool string capacity (ft)"
            ws[cell4] = "Total PCE height (ft)"

        elif column == 'F':
            # --- Detect key tools and compute sectional totals ---
            stuffing_index = None
            bop_index = None
            for idx, row_data in enumerate(data):
                tool_name = str(row_data[2]).lower()
                if any(x in tool_name for x in ["stuffing box", "parasheave"]) and stuffing_index is None:
                    stuffing_index = idx
                if "bop" in tool_name and bop_index is None:
                    bop_index = idx

            # Default to total length if key tools not found
            if stuffing_index is None or bop_index is None or bop_index <= stuffing_index:
                ws[cell1] = total_length
                ws[cell2] = total_length
            else:
                # Compute total length between Stuffing Box/Parasheave and BOP
                length_between = 0.0
                for row_data in data[stuffing_index + 1:bop_index]:
                    length_between += get_number(row_data[5])

                # Compute total length below BOP
                length_below = 0.0
                for row_data in data[bop_index + 1:]:
                    length_below += get_number(row_data[5])

                # Compute total length below BOP
                length_capacity = 0.0
                for row_data in data[stuffing_index + 1:]:
                    length_capacity += get_number(row_data[5])

                ws[cell1] = round(length_between, 2)
                ws[cell2] = round(length_below, 2)
                ws[cell3] = round(length_capacity, 2)

            ws[cell4] = total_length

        elif column == 'G':
            ws[cell1] = "Total Weight (kg)"
            ws[cell3] = "Total Weight (MT)"

    DATA_START_ROW = 7  # your first data row (headers at 6)
    summary_start_row = last_row - 7  # the row that contains "Max OD..." title

    print(321)
    # 1) Clear borders in the data body (C..H, from DATA_START_ROW up to the row before Max OD...)
    for r in range(DATA_START_ROW, summary_start_row):
        for c in range(3, 10):  # C(3) .. J(8)
            ws.cell(row=r, column=c).border = Border()

    # 2) Vertical-only borders for data rows; add a top border on the first data row and a bottom
    #    border on the last data row in the body (summary_start_row-1)
    for r in range(DATA_START_ROW, summary_start_row):
        for c in range(3, 10):
            top = thin if r == DATA_START_ROW else None
            bottom = thin if r == (summary_start_row - 1) else None
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=top, bottom=bottom
            )

    # 3) Re-apply alignment for header + data body only (won’t touch Remarks)
    #    Header row = 6; data body = DATA_START_ROW .. summary_start_row-1
    for c in range(3, 10):
        ws.cell(row=6, column=c).alignment = Alignment(horizontal="center", vertical="center")

    for r in range(DATA_START_ROW, summary_start_row):
        for c in range(3, 10):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    # === Logo & footer (unchanged) ===
    ws.row_dimensions[1].height = 55
    row_1_height = ws.row_dimensions[1].height
    logo_path = get_icon_path('logo_report')
    logo_img = PILImage.open(logo_path)
    max_logo_height = row_1_height * 1.25
    scale_factor = min(1, max_logo_height / logo_img.height)
    logo_img = logo_img.resize((int(logo_img.width * scale_factor), int(logo_img.height * scale_factor)),
                               PILImage.LANCZOS)
    img_bytes = BytesIO()
    logo_img.save(img_bytes, format="PNG")
    img_bytes.seek(0)
    excel_logo = ExcelImage(img_bytes)
    ws.add_image(excel_logo, "A1")

    print(361)
    # **Page Layout Settings**
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:{last_column}{ws.max_row}"
    # **Save Excel File**
    # === Additional merges, borders, and min ID ===
    # === Additional merges, borders, alignment, and min ID ===
    try:
        thin = Side(style="thin")

        # --- Safe merges ---
        merge_ranges = [
            "A1:I1",
            "C2:E2",
            "C3:E3",
            "F2:G2",
            "F3:G3",
            "H2:I2",
            "H3:I3",
            "A4:I4",
            "A5:I5",
            "C45:C46",
            "C47:C48",
            "D45:D46",
            "D47:D48",
            "G45:H46",
            "G47:H48",
            "I45:I46",
            "I47:I48",
            "C49:I49",
            "C50:I52",
        ]

        for rng in merge_ranges:
            safe_merge(ws, rng)

        # --- Center align title and key cells ---
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D4"].alignment = Alignment(horizontal="center", vertical="center")
        ws["D5"].alignment = Alignment(horizontal="center", vertical="center")

        # --- Apply outer grid borders (A1:J52) ---
        for r in range(1, 53):
            for c in range(1, 10):  # A..J
                cell = ws.cell(row=r, column=c)
                # Skip the main data zone (C7:J44) to preserve its custom borders
                if 7 <= r <= 44 and 3 <= c <= 10:
                    continue
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # --- Compute Minimum ID (remove " and ignore non-numeric) ---
        id_values = []
        for r in range(7, 45):  # D7:D44
            val = ws.cell(row=r, column=4).value
            if val:
                try:
                    if val != 'nan"':
                        numeric_val = float(str(val).replace('"', '').strip())
                        id_values.append(numeric_val)
                except ValueError:
                    pass

        if id_values:
            ws["D45"] = min(id_values)
            ws["D45"].font = Font(bold=True)
            ws["D45"].alignment = Alignment(horizontal="center", vertical="center")

            ws["D47"] = max(id_values)
            ws["D47"].font = Font(bold=True)
            ws["D47"].alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws["D45"] = ""

        ws["I45"] = total_weight
        ws["I45"].font = Font(bold=True)
        ws["I45"].alignment = Alignment(horizontal="center", vertical="center")

        ws["I47"] = round(total_weight/1000, 1)
        ws["I47"].font = Font(bold=True)
        ws["I47"].alignment = Alignment(horizontal="center", vertical="center")

    except Exception as e:
        print(f"⚠️ Merge/border/min-ID patch failed: {e}")


    # === Save Excel File ===
    wb.save(excel_path)

def safe_merge(ws, cell_range):
    """Safely merge only if no overlap with existing merged cells."""
    new_min_col, new_min_row, new_max_col, new_max_row = range_boundaries(cell_range)
    for mcr in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(mcr))
        # Check overlap
        if not (new_max_col < min_col or new_min_col > max_col or
                new_max_row < min_row or new_min_row > max_row):
            print(f"⚠️ Skipping overlapping merge: {cell_range}")
            return
    ws.merge_cells(cell_range)

def export_to_pdf(excel_path, pdf_path):
    """Converts an Excel file to PDF and ensures Excel closes properly."""
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # ✅ **Use absolute paths to prevent directory issues**
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Sheets(1)

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        ws.PageSetup.PaperSize = 9
        ws.PageSetup.Orientation = 1

        # ✅ **Ensure the file is closed before exporting to PDF**
        time.sleep(2)
        
        wb.ExportAsFixedFormat(0, pdf_path)
        print(f"✅ Successfully exported PDF: {pdf_path}")

    except Exception as e:
        print(f"❌ Failed to export PDF: {e}")
    finally:
        try:
            wb.Close(SaveChanges=False)
            excel.Quit()
            print("✅ Excel closed successfully.")
        except:
            print("⚠️ Excel instance may still be running.")

    print(f"Excel Path: {excel_path}")
    print(f"PDF Path: {pdf_path}")
def extract_tool_data(drop_zone):
    """Extracts tool details and images from the drop zone."""
    data = []
    tool_images = []

    for i in range(drop_zone.layout.count()):
        widget = drop_zone.layout.itemAt(i).widget()

        # ✅ Ensure the widget is a ToolWidget
        if not isinstance(widget, ToolWidget):
            continue

        # **Extract Tool Information**
        tool_name = widget.tool_name
        size = widget.nominal_size_selector.currentText()
        id = widget.id_label.text()
        top_connection = widget.top_connection_label.text().strip()
        lower_connection = widget.lower_connection_label.text().strip()
        length = widget.length_label.text()
        weight = get_number(widget.weight_label.text())
        service = widget.service_label.currentText()
        wp = widget.wp_label.text()

        # ✅ Combine Top & Bottom connections in one cell with a line break
        connection_text = f"{top_connection}\n{lower_connection}".strip()

        # **Append tool details to the list**
        if "X-Over" in tool_name:
            data.append([
                "", "", "X-Over", id, connection_text, length, service, wp, weight
            ])
        elif "Parasheave" in tool_name:
            data.append([
                "", "", f"{tool_name}", id, connection_text, length, service, wp, weight
            ])
        else:
            data.append([
                "", "", f"{tool_name} ({size})", id, connection_text, length, service, wp, weight
            ])

        # **Retrieve Tool Image**
        image_path = get_pce_image_path(tool_name)
        pixmap = QPixmap(image_path)

        if widget.image_label.pixmap() and not widget.image_label.pixmap().isNull():
            tool_images.append(pixmap.toImage())

    return data, tool_images

