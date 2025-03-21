import os
import time
import win32com.client
from PyQt6.QtGui import QPixmap
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Border, Side, Alignment
from datetime import datetime
from PIL import Image as PILImage
from editor.logic_image_processing import expand_and_center_images, combine_tool_images
from editor.logic_utils import get_number
from ui.ui_toolwidget import ToolWidget
from utils.get_resource_path import get_resource_path

def export_to_excel(excel_path, pdf_path, client_name, location, well_no, well_type, operation_details, comments, drop_zone):
    """Exports tool string configuration to Excel and PDF."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Tool String"
    last_row = 55
    last_column = 'G'

    well_details = [client_name, location, well_no, well_type, operation_details]

    # ✅ Generate Filename
    toolstring_title = "Tool_String"
    for detail in well_details[1:]:  # Skip client name
        if detail and detail not in ["Oil Producer", "Gas Producer"]:
            toolstring_title += f" - {detail}"
    # ✅ Generate directories and filenames
    current_directory = os.getcwd()
    final_directory = os.path.join(current_directory, toolstring_title)
    os.makedirs(final_directory, exist_ok=True)

    # png_path = os.path.join(final_directory, f"{toolstring_title}.png")
    png_path = excel_path.replace(".xlsx",".png")

    # **Define border style**
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    print('designing the excel layout')
    # **Set title**
    ws["A1"] = "TOOL STRING SCHEMATIC"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A4"] = "Operation Details"
    ws["A4"].font = Font(bold=True)
    ws["A5"] = well_details[4]
    
    today = datetime.today().strftime('%d-%m-%Y')

    # **Client Information Section**
    client_info = [
        ["Client Name", "", "Location", "", "Well No.", "Well Type", "Date"],
        [well_details[0], "", well_details[1], "", well_details[2], well_details[3], today]
    ]

    for row_idx, row_data in enumerate(client_info, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx == 2:
                cell.font = Font(bold=True)

    # **Tool Data**
    data = []
    tool_images = []

    print('getting info from the tool list')
    for i in range(drop_zone.layout.count()):  # ❌ Remove self.
        widget = drop_zone.layout.itemAt(i).widget()
        if isinstance(widget, ToolWidget):
            tool_name = widget.tool_name
            size = widget.nominal_size_selector.currentText()
            od = widget.od_label.text()
            od = get_number(od)  # Ensure numeric value
            lower_connection = widget.connection_label.currentText()
            length = widget.length_label.text()
            weight = widget.weight_label.text()
            weight = get_number(weight)  # Ensure numeric value
            data.append(["", "", f"{tool_name} ({size})", od, lower_connection, length, weight])

            if "X-Over" in tool_name:
                tool_name = "X-Over"
            image_file = f"{tool_name}.png".replace('"', '').replace("'", "")
            image_path = get_resource_path(os.path.join("assets", "images", image_file))
            if os.path.exists(image_path):
                pixmap = QPixmap(image_path)
            else:
                dummy_path = get_resource_path(os.path.join("assets", "images", "Dummy Image.png"))
                pixmap = QPixmap(dummy_path)  # Fallback

            if widget.image_label.pixmap() and not widget.image_label.pixmap().isNull():
                # tool_images.append(widget.image_label.pixmap().toImage())
                tool_images.append(pixmap.toImage())


    # **Table Headers**
    headers = ["Diagram", "", "Description", "OD (inches)", "Lower Connection", "Length (ft)", "Weight (lbs)"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # **Calculate Totals**
    total_length = 0.0
    total_weight = 0.0
    od_list = []

    for row_data in data:
        ws.append(row_data)
        od_list.append(row_data[3])
        total_length += get_number(row_data[5])
        total_weight += row_data[6]

    max_od = max(od_list)

    for column in ['C', 'D', 'E', 'F', 'G']:
        cell1 = column + str(last_row - 7)
        cell2 = column + str(last_row-6)
        ws[cell1].font = Font(bold=True)
        ws[cell2].font = Font(bold=True)
        if column == 'C':
            ws[cell1] = "Max OD (in)"
            ws[cell2] = "Max OD (mm)"
        elif column == 'D':
            ws[cell1] = max_od
            ws[cell2] = round(max_od*25.4,1)
        if column == 'E':
            ws[cell1] = "Total Length (ft) & Weight (lbs)"
            ws[cell2] = "Total Length (m) & Weight (kg)"
        elif column == 'F':
            ws[cell1] = total_length
            ws[cell2] = round(total_length*0.3048,1)
        elif column == 'G':
            ws[cell1] = total_weight
            ws[cell2] = round(total_weight*0.453592,1)

    cell_remarks_title = 'C' + str(last_row - 5)
    cell_remarks_content = 'C' + str(last_row-4)
    ws[cell_remarks_title] = "Remarks"
    ws[cell_remarks_title].font = Font(bold=True)

    # **Apply Borders & Alignment**
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # **Auto-adjust Column Widths**
    for col in ws.iter_cols(min_col=3, max_col=7):
        max_length = 0
        try:
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2
        except:
            pass

    # **Merging Cells**
    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:B2")
    ws.merge_cells("A3:B3")
    ws.merge_cells("A6:B6")
    ws.merge_cells("C2:D2")
    ws.merge_cells("C3:D3")
    ws.merge_cells("A4:G4")
    ws.merge_cells("A5:G5")
    ws.merge_cells(f"A7:B{last_row}")
    ws.merge_cells(f"{cell_remarks_title}:{last_column}{last_row - 5}")
    ws.merge_cells(f"{cell_remarks_content}:{last_column}{last_row}")

    ws[cell_remarks_content] = comments.replace("\n", "\n")  # Ensures new lines are preserved
    # Enable text wrapping and set left alignment
    ws[cell_remarks_content].alignment = Alignment(
        wrapText=True,  # ✅ Allows multi-line text
        horizontal="left",  # ✅ Left-align text
    )

    # **Insert Tool String Image**
    cell = "B8"
    if tool_images:
        print('formating images')
        centered_images = expand_and_center_images(tool_images,50)
        tool_image = combine_tool_images(centered_images)
        tool_image.save(png_path)

        pil_img = PILImage.open(png_path)
        max_height = 900
        if pil_img.height > max_height:
            scale_factor = max_height / pil_img.height
            new_width = int(pil_img.width * scale_factor)
            pil_img = pil_img.resize((new_width, max_height), PILImage.LANCZOS)
            pil_img.save(png_path)

        print('retrieving image')
        img = ExcelImage(png_path)
        print('adding image to excel')
        ws.add_image(img, cell)
        print('image added')

    ws.row_dimensions[1].height = last_row
    ws.column_dimensions['A'].width = 4.22
    ws.column_dimensions['B'].width = 4.22 + img.width / 7
    logo_path = get_resource_path(os.path.join("assets", "images", "logo_report.png"))
    # ✅ Load the logo as an ExcelImage before adding it to the worksheet
    logo_img = ExcelImage(logo_path)
    ws.add_image(logo_img, "A1")

    footer_cell = last_column + str(last_row + 2)
    ws[footer_cell] = "This report was computer generated using Deleum Tool String Editor"
    ws[footer_cell].alignment = Alignment(
        horizontal="right",  # ✅ Left-align text
    )

    # **Page Layout Settings**
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = f"A1:G{ws.max_row}"
    # **Save Excel File**
    print('saving to excel')
    wb.save(excel_path)
    print(f"✅ Excel export successful: {excel_path}")

    # ✅ **Wait 2 seconds before opening with Win32 to prevent file access issues**
    time.sleep(2)

    # ✅ **Convert Excel to PDF**
    export_to_pdf(excel_path, pdf_path)

def export_to_pdf(excel_path, pdf_path):
    """Converts an Excel file to PDF and ensures Excel closes properly."""
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        # ✅ **Use absolute paths to prevent directory issues**
        excel_path = os.path.abspath(excel_path)
        pdf_path = os.path.abspath(pdf_path)

        wb = excel.Workbooks.Open(excel_path)
        ws = wb.Sheets(1)

        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False
        ws.PageSetup.PaperSize = 9
        ws.PageSetup.Orientation = 1

        # ✅ **Ensure the file is closed before exporting to PDF**
        time.sleep(2)
        
        wb.ExportAsFixedFormat(0, pdf_path)
        print(f"✅ Successfully exported PDF: {pdf_path}")

    except Exception as e:
        print(f"❌ Failed to export PDF: {e}")
    finally:
        try:
            wb.Close(SaveChanges=False)
            excel.Quit()
            print("✅ Excel closed successfully.")
        except:
            print("⚠️ Excel instance may still be running.")

    print(f"Excel Path: {excel_path}")
    print(f"PDF Path: {pdf_path}")
