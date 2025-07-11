import datetime
import io
import os
import re
import shutil
import tempfile

import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import xlrd
import matplotlib.dates as mdates
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.widgets import SpanSelector
from PyQt6.QtCore import Qt, QTimer, QSize
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QImage, QIcon
from PyQt6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QPushButton, QFileDialog,
    QMessageBox, QTableWidget, QTableWidgetItem, QHBoxLayout, QApplication,
    QSplitter, QGridLayout, QGroupBox, QStackedWidget, QTimeEdit, QLineEdit, QToolButton
)

# Local imports
from ui.components.ui_footer import FooterWidget
from ui.components.ui_sidebar_widget import SidebarWidget
from ui.components.ui_titlebar import CustomTitleBar
from ui.windows.ui_messagebox_window import MessageBoxWindow
from utils.path_finder import get_icon_path, get_path
from utils.styles import GROUPBOX_STYLE, MESSAGEBOX_STYLE
from utils.theme_manager import apply_theme, toggle_theme


class QuadrupleDragDropWidget(QWidget):
    def __init__(self, callback, parent=None):
        super().__init__(parent)
        self.callback = callback
        self.setAcceptDrops(True)

        # Define file type configurations
        self.file_types = {
            "top": {
                "label": "Drag & Drop Top Gauge Data File (.txt)",
                "drop_text": "Top Gauge Data\n(.txt)",
                "extensions": ('.txt',),
                "file_path": None
            },
            "bottom": {
                "label": "Drag & Drop Bottom Gauge Data File (.txt)",
                "drop_text": "Bottom Gauge Data\n(.txt)",
                "extensions": ('.txt',),
                "file_path": None
            },
            "timesheet": {
                "label": "Drag & Drop Survey Timesheet (.xls/.xlsx)",
                "drop_text": "Survey Timesheet\n(.xls/.xlsx)",
                "extensions": ('.xls', '.xlsx', '.XLS'),
                "file_path": None
            },
            "tvd": {
                "label": "Drag & Drop TVD Calculation File (.xls/.xlsx)",
                "drop_text": "TVD Calculation\n(.xls/.xlsx)",
                "extensions": ('.xls', '.xlsx', '.XLS'),
                "file_path": None
            }
        }

        layout = QGridLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)

        # Create drop areas and store references
        self.drop_areas = {}
        self.labels = {}
        self.clear_buttons = {}
        self.browse_buttons = {}

        for i, (file_type, config) in enumerate(self.file_types.items()):
            row = i // 2
            col = i % 2

            # Create UI components
            label = QLabel(config["label"])
            drop_area = self.create_drop_area(config["drop_text"])
            buttons_layout = self.create_file_buttons(file_type)

            # Create group box
            title = file_type.upper() + (" GAUGE" if file_type in ["top", "bottom"] else "")
            group_box = self.create_group_box(title, drop_area, label, buttons_layout)

            # Add to layout and store references
            layout.addWidget(group_box, row, col)
            self.drop_areas[file_type] = drop_area
            self.labels[file_type] = label

        # Process and Clear buttons
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)

        self.clear_all_btn = self.create_button("Clear All", "#e74c3c", "#c0392b", 40)
        self.clear_all_btn.clicked.connect(self.clear_all_files)
        button_layout.addWidget(self.clear_all_btn)

        self.process_btn = self.create_button("Process Files", "#3498db", None, 40)
        self.process_btn.setEnabled(False)
        self.process_btn.clicked.connect(self.process_files)
        button_layout.addWidget(self.process_btn)

        layout.addWidget(button_container, 2, 0, 1, 2)

    # --- Helper Methods ---#
    def create_button(self, text, color, hover_color=None, height=None):
        """Create standardized buttons with consistent styling"""
        button = QPushButton(text)
        if height:
            button.setMinimumHeight(height)

        style = f"""
            QPushButton {{
                background-color: {color};
                color: white;
                font-weight: bold;
                border-radius: 5px;
                padding: 8px;
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
                color: #888888;
            }}
        """

        if hover_color:
            style += f"""
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            """

        button.setStyleSheet(style)
        return button

    def create_drop_area(self, text):
        """Create a consistent drop area widget"""
        drop_area = QLabel(text)
        drop_area.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px dashed #aaa;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(240, 240, 240, 100);
            }
        """)
        drop_area.setMinimumSize(200, 180)
        drop_area.setAcceptDrops(True)
        drop_area.dragEnterEvent = self.dragEnterEvent
        drop_area.dropEvent = lambda event, da=drop_area: self.dropEvent(event, da)
        return drop_area

    def create_file_buttons(self, file_type):
        """Create browse/clear buttons for a file type"""
        button_layout = QHBoxLayout()

        browse_btn = self.create_button("Browse", "#f0f0f0", "#e0e0e0", 25)
        browse_btn.setStyleSheet(browse_btn.styleSheet() + "color: black;")
        browse_btn.clicked.connect(lambda _, ft=file_type: self.browse_file(ft))

        clear_btn = self.create_button("Clear", "#f8d7da", "#f5c6cb", 25)
        clear_btn.setStyleSheet(clear_btn.styleSheet() + "color: #721c24;")
        clear_btn.setEnabled(False)
        clear_btn.clicked.connect(lambda _, ft=file_type: self.clear_file(ft))

        button_layout.addWidget(browse_btn)
        button_layout.addWidget(clear_btn)

        # Store button references
        self.browse_buttons[file_type] = browse_btn
        self.clear_buttons[file_type] = clear_btn

        return button_layout

    def create_group_box(self, title, widget, label, buttons):
        """Create a standardized group box container"""
        group = QGroupBox(title)
        group_layout = QVBoxLayout(group)
        group_layout.addWidget(label)
        group_layout.addWidget(widget)
        group_layout.addLayout(buttons)
        group.setStyleSheet(GROUPBOX_STYLE)
        return group

    # --- File Handling ---#
    def handle_file_drop(self, drop_area, file_path):
        """Handle file drop on a specific area"""
        for file_type, config in self.file_types.items():
            if drop_area == self.drop_areas[file_type]:
                if file_path.lower().endswith(config["extensions"]):
                    self.set_file(file_type, file_path)
                    return

        MessageBoxWindow.message_simple(self, "Invalid File", "Please drop the correct file type in each area")

    def browse_file(self, file_type):
        """Open file dialog for specific file type"""
        config = self.file_types[file_type]
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Open {file_type.capitalize()} File",
            "",
            f"Files ({' '.join(['*' + ext for ext in config['extensions']])})"
        )
        if file_path:
            self.set_file(file_type, file_path)

    def set_file(self, file_type, file_path):
        """Set file for a specific type and update UI"""
        config = self.file_types[file_type]
        config["file_path"] = file_path
        filename = file_path.split('/')[-1]

        drop_area = self.drop_areas[file_type]
        drop_area.setText(f"Loaded:\n{filename}")
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px solid #2ecc71;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(200, 255, 200, 150);
            }
        """)

        self.clear_buttons[file_type].setEnabled(True)
        self.update_process_button()

    def clear_file(self, file_type):
        """Clear file for a specific type and reset UI"""
        config = self.file_types[file_type]
        config["file_path"] = None

        drop_area = self.drop_areas[file_type]
        drop_area.setText(config["drop_text"])
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px dashed #aaa;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(240, 240, 240, 100);
            }
        """)

        self.clear_buttons[file_type].setEnabled(False)
        self.update_process_button()

    def clear_all_files(self):
        """Clear all uploaded files and reset UI"""
        for file_type in self.file_types:
            self.clear_file(file_type)

    def update_process_button(self):
        """Enable process button only when all files are loaded"""
        all_loaded = all(
            config["file_path"] is not None
            for config in self.file_types.values()
        )
        self.process_btn.setEnabled(all_loaded)

    # --- Event Handlers ---#
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            widget = self.childAt(event.position().toPoint())
            if widget in self.drop_areas.values():
                widget.setStyleSheet("""
                    QLabel {
                        border: 3px dashed #3498db;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(200, 230, 255, 150);
                    }
                """)

    def dropEvent(self, event: QDropEvent, drop_area):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_path = urls[0].toLocalFile()
            self.handle_file_drop(drop_area, file_path)
            # Reset style after drop
            if self.file_types[[k for k, v in self.drop_areas.items() if v == drop_area][0]]["file_path"]:
                drop_area.setStyleSheet("""
                    QLabel {
                        border: 3px solid #2ecc71;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(200, 255, 200, 150);
                    }
                """)
            else:
                drop_area.setStyleSheet("""
                    QLabel {
                        border: 3px dashed #aaa;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(240, 240, 240, 100);
                    }
                """)

    def process_files(self):
        """Process files when all are loaded"""
        if all(config["file_path"] for config in self.file_types.values()):
            self.callback(
                self.file_types["top"]["file_path"],
                self.file_types["bottom"]["file_path"],
                self.file_types["timesheet"]["file_path"],
                self.file_types["tvd"]["file_path"]
            )


class SGSTXTApp(QWidget):
    def __init__(self):
        super().__init__()
        self.location = None
        self.well = None
        self.setWindowTitle("SGS / FGS txt processing")
        self.setMinimumSize(1400, 800)
        self.current_theme = "Deleum"
        self.init_ui()
        self.sidebar_expanded = False
        self.station_timings = []
        self.top_data = []
        self.bottom_data = []
        self.top_file_path = None
        self.bottom_file_path = None
        self.timesheet_file_path = None
        self.ahd_tvd_map = {}
        self.events = []
        self.current_canvas = None
        self.bdf = 0
        self.sea_level = 0

    def init_ui(self):
        main_container = QVBoxLayout(self)
        main_container.setContentsMargins(0, 0, 0, 0)
        main_container.setSpacing(0)

        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)

        items = [
            (get_icon_path('save'), "Save Survey", self.save_file, "Save the current survey (Ctrl+S)"),
            (get_icon_path('load'), "Load Survey", lambda: self.open_file_dialog('top'), "Open a survey (Ctrl+O)"),
            (get_icon_path('export'), "Process && Export", self.process_data, "Export to Interpretation File")
        ]

        self.sidebar = SidebarWidget(self, items)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.title_bar = CustomTitleBar(self, lambda: self.sidebar.toggle_visibility(), "SGS/FGS txt processing")

        main_container.addWidget(self.title_bar)
        main_layout.addWidget(self.sidebar)

        # Main content area - using stacked widget for screens
        self.content_stack = QStackedWidget()
        main_layout.addWidget(self.content_stack)

        # File upload screen
        self.file_upload_widget = QWidget()
        file_upload_layout = QVBoxLayout(self.file_upload_widget)
        file_upload_layout.setContentsMargins(20, 20, 20, 20)

        # Triple drag and drop area
        self.drag_drop_widget = QuadrupleDragDropWidget(self.process_all_files, self)
        file_upload_layout.addWidget(self.drag_drop_widget, 1)

        # Footer
        self.footer = FooterWidget(self, theme_callback=self.toggle_theme)
        self.theme_button = self.footer.theme_button
        file_upload_layout.addWidget(self.footer)

        self.content_stack.addWidget(self.file_upload_widget)

        # Results display screen
        self.results_widget = QWidget()
        results_layout = QVBoxLayout(self.results_widget)
        results_layout.setContentsMargins(20, 20, 20, 20)

        # Navigation bar at top of results
        nav_bar = QWidget()
        nav_layout = QHBoxLayout(nav_bar)
        nav_layout.setContentsMargins(0, 0, 0, 10)

        self.back_button = QPushButton("← Back to File Upload")
        self.back_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.back_button.setFixedHeight(40)
        self.back_button.clicked.connect(self.show_file_upload)
        nav_layout.addWidget(self.back_button)

        results_layout.addWidget(nav_bar)

        # Results content
        results_content = QWidget()
        results_content_layout = QHBoxLayout(results_content)
        results_content_layout.setContentsMargins(0, 0, 0, 0)

        # Graph area - will contain two stacked graphs
        self.graph_widget = QWidget()
        self.graph_layout = QVBoxLayout(self.graph_widget)
        self.graph_layout.setContentsMargins(0, 0, 0, 0)

        # Table area - now in a group box
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)

        # Create group box for the table
        table_group = QGroupBox("Station Data")
        table_group.setStyleSheet(GROUPBOX_STYLE)
        table_group_layout = QVBoxLayout(table_group)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(19)
        self.table_widget.setHorizontalHeaderLabels([
            "Station", "Depth (ft)", "Start Time", "End Time", "Duration (min)",
            "AHD FTBDF", "TVD FTBDF",
            "High P (T)", "Low P (T)", "Med P (T)", "High T (T)", "Low T (T)", "Med T (T)",
            "High P (B)", "Low P (B)", "Med P (B)", "High T (B)", "Low T (B)", "Med T (B)"
        ])
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #f8f8f8;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 4px;
                border: 1px solid #ccc;
                font-weight: bold;
            }
        """)

        # Add table to group
        table_group_layout.addWidget(self.table_widget)

        # Create button row with two group boxes
        button_row = QWidget()
        button_layout = QHBoxLayout(button_row)
        button_layout.setContentsMargins(0, 10, 0, 0)
        button_layout.setSpacing(20)  # Add spacing between group boxes

        # Group Box 1: Template Downloads
        template_group = QGroupBox("Download Blank Templates")
        template_group.setStyleSheet(GROUPBOX_STYLE)
        template_layout = QVBoxLayout(template_group)
        template_layout.setSpacing(8)  # Add spacing between buttons

        self.download_template_button = QPushButton("Interpretation")
        self.download_template_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.download_template_button.clicked.connect(self.download_interpretation_template)

        self.download_md_tvd_button = QPushButton("TVD Calculation")
        self.download_md_tvd_button.setStyleSheet("""
            QPushButton {
                background-color: #17a2b8;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #138496;
            }
        """)
        self.download_md_tvd_button.clicked.connect(self.download_md_tvd_template)

        template_layout.addWidget(self.download_template_button)
        template_layout.addWidget(self.download_md_tvd_button)

        # Group Box 2: Actions
        action_group = QGroupBox("Actions")
        action_group.setStyleSheet(GROUPBOX_STYLE)
        action_layout = QVBoxLayout(action_group)
        action_layout.setSpacing(8)

        self.copy_button = QPushButton()
        self.copy_button.setIcon(QIcon(get_icon_path('copy')))  # Add copy icon
        self.copy_button.setText("Copy Statistics")
        self.copy_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.copy_button.clicked.connect(self.copy_statistics)
        self.copy_button.setEnabled(False)  # Disabled until data is loaded

        self.generate_as2_button = QPushButton("Generate AS2 Files")
        self.generate_as2_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.generate_as2_button.clicked.connect(self.generate_as2_files)
        self.generate_as2_button.setEnabled(False)  # Disabled until data is loaded

        self.process_data_button = QToolButton()
        self.process_data_button.setIcon(QIcon(get_icon_path('export')))
        self.process_data_button.setText("Process && Export")
        self.process_data_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon
)
        self.process_data_button.setIconSize(QSize(32, 32))  # Increase icon size
        self.process_data_button.setStyleSheet("""
            QToolButton {
                background-color: #28a745;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QToolButton:hover {
                background-color: #218838;
            }
            QToolButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.process_data_button.clicked.connect(self.process_data)
        self.process_data_button.setEnabled(False)  # Disabled until data is loaded

        action_layout.addWidget(self.copy_button)
        action_layout.addWidget(self.generate_as2_button)

        # Add group boxes to the button layout
        button_layout.addWidget(template_group, 1)
        button_layout.addWidget(action_group, 1)
        button_layout.addWidget(self.process_data_button)

        # Create event section
        event_group = QGroupBox("Events")
        event_group.setStyleSheet(GROUPBOX_STYLE)
        event_layout = QGridLayout(event_group)

        # Event input fields
        self.event_time_edit = QTimeEdit()
        self.event_time_edit.setDisplayFormat("HH:mm:ss")

        self.event_desc_edit = QLineEdit()
        self.event_desc_edit.setPlaceholderText("Enter event description")

        self.add_event_btn = QPushButton("Add Event")
        self.add_event_btn.clicked.connect(self.add_event)
        self.add_event_btn.setStyleSheet("""
            QPushButton {
                background-color: #5cb85c;
                color: white;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #4cae4c;
            }
        """)

        self.remove_event_btn = QPushButton("Remove Selected")
        self.remove_event_btn.clicked.connect(self.remove_event)
        self.remove_event_btn.setStyleSheet("""
            QPushButton {
                background-color: #d9534f;
                color: white;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #c9302c;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)

        # Event table
        self.event_table = QTableWidget(0, 2)
        self.event_table.setHorizontalHeaderLabels(["Time", "Event Description"])
        self.event_table.horizontalHeader().setStretchLastSection(True)
        self.event_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # Add widgets to layout
        event_layout.addWidget(QLabel("Time:"), 0, 0)
        event_layout.addWidget(self.event_time_edit, 0, 1)
        event_layout.addWidget(QLabel("Description:"), 0, 2)
        event_layout.addWidget(self.event_desc_edit, 0, 3)
        event_layout.addWidget(self.add_event_btn, 0, 4)
        event_layout.addWidget(self.remove_event_btn, 0, 5)
        event_layout.addWidget(self.event_table, 1, 0, 1, 6)

        # Add event section to table container
        table_layout.addWidget(event_group)
        table_layout.addWidget(button_row)

        # Inside the button_row section for results screen:
        self.copy_graphs_button = QPushButton("Copy Graphs")
        self.copy_graphs_button.setFixedWidth(150)
        self.copy_graphs_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.copy_graphs_button.clicked.connect(self.copy_graphs)
        self.copy_graphs_button.setEnabled(False)  # Disabled until data is loaded

        # Add group to container
        table_layout.addWidget(event_group)
        table_layout.addWidget(table_group)  # Now using group box
        table_layout.addWidget(button_row)

        # Add widgets to splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.graph_widget)
        splitter.addWidget(table_container)
        splitter.setSizes([500, 500])
        results_content_layout.addWidget(splitter)

        results_layout.addWidget(results_content, 1)

        # Footer for results screen
        results_footer = FooterWidget(self, theme_callback=self.toggle_theme)
        results_layout.addWidget(results_footer)

        self.content_stack.addWidget(self.results_widget)

        main_container.addLayout(main_layout)
        self.setLayout(main_container)

        # Apply theme after UI is set up
        apply_theme(self, self.current_theme)

    def process_sgs_txt_file(self, file_path):
        try:
            with open(file_path, 'r') as f:
                lines = f.readlines()

            # Find where data starts
            data_start = 0
            for i, line in enumerate(lines):
                if "Date" in line and "Time" in line and "Press" in line:
                    data_start = i + 2  # Skip header and unit line
                    break

            if data_start == 0:
                QMessageBox.critical(self, "Error", "Could not find data headers in file")
                return None

            # Parse data
            data = []
            for line in lines[data_start:]:
                parts = line.strip().split()
                if len(parts) < 4:
                    continue

                try:
                    date_str = parts[0].replace("-", "/")
                    time_str = parts[1]
                    date_time = datetime.datetime.strptime(
                        f"{date_str} {time_str}", "%d/%m/%Y %H:%M:%S"
                    )
                    pressure = float(parts[2])
                    temperature = float(parts[3])
                    data.append((date_time, pressure, temperature))
                except ValueError:
                    continue

            return data if data else None

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process data file:\n{str(e)}")
            return None

    def process_excel_timesheet(self, file_path):
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, header=None, engine='openpyxl')
            elif lower_path.endswith('.xls'):  # Now handles both .xls and .XLS
                df = pd.read_excel(file_path, header=None, engine='xlrd')
            else:
                QMessageBox.critical(self, "Error", "Unsupported timesheet format. Must be .xlsx or .xls")
                return []
            # Find the date - handles both formats
            date_value = None
            for i in range(len(df)):
                cell_value = str(df.iloc[i, 2])
                if "Date of Survey" in cell_value or "Date :" in cell_value:
                    # Main format: "Date of Survey  :  1/05/2025"
                    if "Date of Survey" in cell_value:
                        date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', cell_value)
                        if date_match:
                            date_value = datetime.datetime.strptime(date_match.group(1), "%d/%m/%Y").date()
                    # Alternate format: "Date : 01.05.2025"
                    elif "Date :" in cell_value:
                        date_match = re.search(r'Date :\s*(\d{2}\.\d{2}\.\d{4})', cell_value)
                        if date_match:
                            date_value = datetime.datetime.strptime(date_match.group(1), "%d.%m.%Y").date()
                    break

            if date_value is None:
                QMessageBox.critical(self, "Error", "Could not find survey date in timesheet")
                return []

            # Determine format and find data start
            format_type = None
            data_start = None

            # Check for main format (has "No" in D and "Depth (ft)" in E)
            for i in range(len(df)):
                col_d = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""
                col_e = str(df.iloc[i, 4]).strip() if pd.notna(df.iloc[i, 4]) else ""

                if col_d == 'No' and col_e == 'Depth (ft)':
                    format_type = 'main'
                    data_start = i + 1  # Data starts next row
                    break

            # Check for alternate format (has "ATM" in D)
            if format_type is None:
                for i in range(len(df)):
                    col_d = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""
                    if col_d == 'ATM':
                        format_type = 'alternate'
                        data_start = i
                        break

            if format_type is None or data_start is None:
                QMessageBox.critical(self, "Error", "Could not determine timesheet format")
                return []

            # Parse station data based on format
            stations = []
            i = data_start

            while i < len(df):
                # Main format columns
                if format_type == 'main':
                    station_type = str(df.iloc[i, 4]).strip() if pd.notna(df.iloc[i, 4]) else ""  # Column E
                    depth = station_type  # Depth is in same column
                    duration = df.iloc[i, 5]  # Column F
                    start_time = df.iloc[i, 6]  # Column G
                    end_time = df.iloc[i, 8]  # Column I

                    # Skip ATM stations
                    if station_type == 'ATM':
                        i += 1
                        continue

                    # Set THP depth to 0
                    if station_type == 'THP':
                        depth = 0

                # Alternate format columns
                else:  # format_type == 'alternate'
                    station_type = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""  # Column D
                    depth = df.iloc[i, 4]  # Column E
                    duration = df.iloc[i, 5]  # Column F
                    start_time = df.iloc[i, 6]  # Column G
                    end_time = df.iloc[i, 8]  # Column I

                    # Skip ATM stations
                    if station_type == 'ATM':
                        i += 1
                        continue

                    # Set THP depth to 0
                    if station_type == 'THP':
                        depth = 0

                # Skip if any critical value is missing
                if (not station_type or
                        pd.isna(duration) or
                        pd.isna(start_time) or
                        pd.isna(end_time)):
                    i += 1
                    continue

                # Convert times to time objects
                start_time_obj = self.parse_time(start_time)
                end_time_obj = self.parse_time(end_time)

                # Create datetime objects
                try:
                    start_dt = datetime.datetime.combine(date_value, start_time_obj)
                    end_dt = datetime.datetime.combine(date_value, end_time_obj)
                except Exception as e:
                    print(f"Error creating datetime: {e}")
                    i += 1
                    continue

                stations.append({
                    'station': station_type,
                    'depth': depth,
                    'start': start_dt,
                    'end': end_dt,
                    'duration': duration
                })
                i += 1

            return stations

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process timesheet:\n{str(e)}")
            return []

    def parse_time(self, time_val):
        """Parse time values from various formats into time objects"""
        if isinstance(time_val, datetime.time):
            return time_val

        if isinstance(time_val, str):
            time_val = time_val.strip()
            # Handle colon-separated times (12:17:00)
            if ':' in time_val:
                parts = time_val.split(':')
                hour = int(parts[0])
                minute = int(parts[1])
                second = int(parts[2]) if len(parts) >= 3 else 0
                return datetime.time(hour, minute, second)
            # Handle 4-digit times (1217)
            else:
                # Remove non-digits and pad with zeros
                digits = ''.join(filter(str.isdigit, time_val)).zfill(4)[:4]
                hour = int(digits[:2])
                minute = int(digits[2:4])
                return datetime.time(hour, minute)

        elif isinstance(time_val, (int, float)):
            # Handle integer times (1217)
            time_val = int(time_val)
            hour = time_val // 100
            minute = time_val % 100
            return datetime.time(hour, minute)

        # Return midnight as default
        return datetime.time(0, 0)

    def show_file_upload(self):
        # Switch back to file upload screen
        self.content_stack.setCurrentIndex(0)

    def init_span_selectors(self, ax_top, ax_bottom):
        """Initialize span selectors for both top and bottom graphs"""
        # Create SpanSelector for top graph
        self.span_selector_top = SpanSelector(
            ax_top,
            self.on_horizontal_select,
            'horizontal',
            useblit=True,
            props=dict(alpha=0.5, facecolor='tab:blue')
        )

        # Create SpanSelector for bottom graph
        self.span_selector_bottom = SpanSelector(
            ax_bottom,
            self.on_horizontal_select,
            'horizontal',
            useblit=True,
            props=dict(alpha=0.5, facecolor='tab:blue')
        )

    def show_stacked_graphs(self, top_file_path, bottom_file_path):
        # Clear previous graph
        while self.graph_layout.count():
            child = self.graph_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        # Create figure with two subplots and a status bar area
        fig = plt.figure(figsize=(10, 8.5))
        gs = fig.add_gridspec(3, 1, height_ratios=[1, 1, 0.1])
        ax_top = fig.add_subplot(gs[0])
        ax_bottom = fig.add_subplot(gs[1], sharex=ax_top)
        ax_status = fig.add_subplot(gs[2])
        ax_status.axis('off')  # Hide axes for status area

        # Store initial x-axis limits for reset functionality
        if self.top_data or self.bottom_data:
            all_times = []
            if self.top_data:
                all_times.extend([item[0] for item in self.top_data])
            if self.bottom_data:
                all_times.extend([item[0] for item in self.bottom_data])
            min_time = min(all_times)
            max_time = max(all_times)
            self.initial_x_lim = (min_time, max_time)
        else:
            self.initial_x_lim = None

        # Plot top gauge data
        if self.top_data:
            times_top = [item[0] for item in self.top_data]
            pressures_top = [item[1] for item in self.top_data]
            temps_top = [item[2] for item in self.top_data]

            # Plot pressure
            ax_top.plot(times_top, pressures_top, 'b-', label='Pressure (psia)', linewidth=1.5)
            ax_top.set_ylabel('Pressure (psia)', color='b', fontsize=10)
            ax_top.tick_params(axis='y', labelcolor='b')
            ax_top.grid(True, linestyle='--', alpha=0.7)
            ax_top.set_title(f"Top Gauge - {top_file_path.split('/')[-1]}", fontsize=10)

            # Add temperature on secondary axis
            ax_top2 = ax_top.twinx()
            ax_top2.plot(times_top, temps_top, 'r-', label='Temperature (°F)', linewidth=1.5)
            ax_top2.set_ylabel('Temperature (°F)', color='r', fontsize=10)
            ax_top2.tick_params(axis='y', labelcolor='r')

        # Plot bottom gauge data
        if self.bottom_data:
            times_bottom = [item[0] for item in self.bottom_data]
            pressures_bottom = [item[1] for item in self.bottom_data]
            temps_bottom = [item[2] for item in self.bottom_data]

            # Plot pressure
            ax_bottom.plot(times_bottom, pressures_bottom, 'b-', label='Pressure (psia)', linewidth=1.5)
            ax_bottom.set_ylabel('Pressure (psia)', color='b', fontsize=10)
            ax_bottom.tick_params(axis='y', labelcolor='b')
            ax_bottom.grid(True, linestyle='--', alpha=0.7)
            ax_bottom.set_title(f"Bottom Gauge - {bottom_file_path.split('/')[-1]}", fontsize=10)

            # Add temperature on secondary axis
            ax_bottom2 = ax_bottom.twinx()
            ax_bottom2.plot(times_bottom, temps_bottom, 'r-', label='Temperature (°F)', linewidth=1.5)
            ax_bottom2.set_ylabel('Temperature (°F)', color='r', fontsize=10)
            ax_bottom2.tick_params(axis='y', labelcolor='r')

        # Format x-axis for both plots
        ax_bottom.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
        ax_bottom.xaxis.set_major_locator(mdates.MinuteLocator(interval=15))
        fig.autofmt_xdate()

        # Create status text
        self.status_text = ax_status.text(
            0.5, 0.5, "",
            ha="center", va="center",
            fontsize=10,
            transform=ax_status.transAxes
        )

        # Initialize cursor lines
        init_x = self.top_data[0][0] if self.top_data else self.bottom_data[0][0] if self.bottom_data else 0
        self.cursor_vline_top = ax_top.axvline(
            x=init_x, color='k', alpha=0.5, linewidth=1, visible=False
        )
        self.cursor_vline_bottom = ax_bottom.axvline(
            x=init_x, color='k', alpha=0.5, linewidth=1, visible=False
        )

        # Initialize markers for data points
        self.top_p_marker, = ax_top.plot([init_x], [0], 'bo', markersize=6, visible=False)
        self.top_t_marker, = ax_top2.plot([init_x], [0], 'ro', markersize=6, visible=False)
        self.bottom_p_marker, = ax_bottom.plot([init_x], [0], 'bo', markersize=6, visible=False)
        self.bottom_t_marker, = ax_bottom2.plot([init_x], [0], 'ro', markersize=6, visible=False)

        # Store data for quick lookup
        self.top_times = mdates.date2num(times_top) if self.top_data else None
        self.bottom_times = mdates.date2num(times_bottom) if self.bottom_data else None
        self.top_pressures = pressures_top if self.top_data else None
        self.top_temps = temps_top if self.top_data else None
        self.bottom_pressures = pressures_bottom if self.bottom_data else None
        self.bottom_temps = temps_bottom if self.bottom_data else None

        # Highlight station timings on both graphs
        colors = plt.cm.tab20(np.linspace(0, 1, len(self.station_timings)))
        for idx, station in enumerate(self.station_timings):
            ax_top.axvspan(
                station['start'], station['end'],
                alpha=0.2, color=colors[idx],
                label=f"{station['station']} - {station['depth']}ft"
            )
            ax_bottom.axvspan(
                station['start'], station['end'],
                alpha=0.2, color=colors[idx],
                label=f"{station['station']} - {station['depth']}ft"
            )

        # Add legend to top plot only
        lines_top, labels_top = ax_top.get_legend_handles_labels()
        lines_top2, labels_top2 = ax_top2.get_legend_handles_labels()
        ax_top.legend(lines_top + lines_top2, labels_top + labels_top2, loc='upper left', fontsize=8)

        fig.tight_layout()

        # Initialize span selectors for both graphs
        self.init_span_selectors(ax_top, ax_bottom)

        # Initialize zoom toolbar
        self.init_zoom_toolbar()

        # Add canvas and toolbar to layout
        canvas = FigureCanvas(fig)
        self.graph_layout.addWidget(canvas)
        self.graph_layout.addWidget(self.toolbar_widget)  # Add zoom toolbar
        self.current_canvas = canvas
        canvas.mpl_connect('motion_notify_event', self.on_mouse_move)

    def populate_station_table(self):
        self.table_widget.setRowCount(len(self.station_timings))

        # Precompute data arrays for vectorized operations
        top_data = np.array(self.top_data, dtype=object) if self.top_data else None
        bottom_data = np.array(self.bottom_data, dtype=object) if self.bottom_data else None

        for row, station in enumerate(self.station_timings):
            # Basic station info
            self.table_widget.setItem(row, 0, QTableWidgetItem(station['station']))
            self.table_widget.setItem(row, 1, QTableWidgetItem(str(station.get('depth', 'N/A'))))
            self.table_widget.setItem(row, 2, QTableWidgetItem(station['start'].strftime("%H:%M:%S")))
            self.table_widget.setItem(row, 3, QTableWidgetItem(station['end'].strftime("%H:%M:%S")))
            self.table_widget.setItem(row, 4, QTableWidgetItem(str(station['duration'])))

            # TVD data
            ahd = self.tvd_data.get('ahd_values', [])
            tvd = self.tvd_data.get('tvd_values', [])
            self.table_widget.setItem(row, 5, QTableWidgetItem(f"{ahd[row]:.2f}" if row < len(ahd) else "N/A"))
            self.table_widget.setItem(row, 6, QTableWidgetItem(f"{tvd[row]:.2f}" if row < len(tvd) else "N/A"))

            # Process gauge statistics
            for col_offset, data, gauge_name in [
                (7, top_data, 'top'),
                (13, bottom_data, 'bottom')
            ]:
                if data is None:
                    continue

                times = data[:, 0].astype('datetime64[us]')
                pressures = data[:, 1].astype(float)
                temps = data[:, 2].astype(float)

                start_dt = np.datetime64(station['start'])
                end_dt = np.datetime64(station['end'])

                mask = (times >= start_dt) & (times <= end_dt)
                if not np.any(mask):
                    stats = ["N/A"] * 6
                else:
                    p_slice = pressures[mask]
                    t_slice = temps[mask]
                    stats = [
                        np.max(p_slice), np.min(p_slice), np.median(p_slice),
                        np.max(t_slice), np.min(t_slice), np.median(t_slice)
                    ]

                for i, stat in enumerate(stats):
                    item = QTableWidgetItem(f"{stat:.2f}" if isinstance(stat, float) else str(stat))
                    self.table_widget.setItem(row, col_offset + i, item)

        # Enable UI components
        for btn in [self.copy_button, self.generate_as2_button, self.copy_graphs_button, self.process_data_button]:
            btn.setEnabled(True)

        self.table_widget.resizeColumnsToContents()

    def set_button_success_feedback(self, button, success_text, success_icon, original_text, original_icon):
        """Set button to success state and schedule reset"""
        button.setStyleSheet(f"""
            QPushButton {{
                background-color: #28a745;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }}
        """)
        button.setIcon(QIcon(get_icon_path(success_icon)))
        button.setText(success_text)

        # Set timer to revert button after 3 seconds
        QTimer.singleShot(3000, lambda: self.reset_button(button, original_text, original_icon))

    def reset_button(self, button, original_text, original_icon):
        """Revert button to original state"""
        # Determine the appropriate style based on button type
        if button == self.copy_button:
            style = """
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border-radius: 5px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #888888;
                }
            """
        elif button == self.copy_graphs_button:
            style = """
                QPushButton {
                    background-color: #9b59b6;
                    color: white;
                    border-radius: 5px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #8e44ad;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #888888;
                }
            """
        else:
            style = ""

        button.setStyleSheet(style)
        button.setIcon(QIcon(get_icon_path(original_icon)))
        button.setText(original_text)

    # Update the copy_statistics method
    def copy_statistics(self):
        """Copy pressure and temperature statistics to clipboard with visual feedback"""
        clipboard = QApplication.clipboard()

        # Get statistics data
        stats_data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for col in range(5, 19):  # Columns 5 to 18
                item = self.table_widget.item(row, col)
                row_data.append(item.text() if item else "")
            stats_data.append("\t".join(row_data))

        # Format as tab-separated values
        text_data = "\n".join(stats_data)

        # Set to clipboard
        clipboard.setText(text_data)

        # Show visual feedback
        self.set_button_success_feedback(
            self.copy_button,
            "Copied!",
            'check',
            "Copy Statistics",
            'copy'
        )

    # Update the copy_graphs method
    def copy_graphs(self):
        """Copy the current graphs to clipboard as an image"""
        if not self.current_canvas:  # Check if canvas exists
            QMessageBox.warning(self, "No Graphs", "No graphs available to copy")
            return

        try:
            # Create a buffer to save the image
            buf = io.BytesIO()
            self.current_canvas.figure.savefig(buf, format='png', dpi=100)
            buf.seek(0)

            # Create QImage from buffer
            image = QImage()
            image.loadFromData(buf.getvalue(), 'PNG')

            # Copy to clipboard
            clipboard = QApplication.clipboard()
            clipboard.setImage(image)

            # Show visual feedback
            self.set_button_success_feedback(
                self.copy_graphs_button,
                "Copied!",
                'check',
                "Copy Graphs",
                'copy'
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to copy graphs:\n{str(e)}")

    def toggle_theme(self):
        self.current_theme = toggle_theme(
            widget=self,
            current_theme=self.current_theme,
            theme_button=self.theme_button,
            summary_widget=None
        )

    # Optimized template download methods
    def _download_template(self, template_name, dialog_title, silent=False):
        """Generic template download handler"""
        try:
            template_path = get_path(f"assets/resources/{template_name}")

            if silent:
                temp_path = os.path.join(tempfile.gettempdir(), os.path.basename(template_name))
                shutil.copy(template_path, temp_path)
                return temp_path

            file_path, _ = QFileDialog.getSaveFileName(
                self, dialog_title, os.path.basename(template_name), "Excel Files (*.xlsx)"
            )

            if file_path:
                shutil.copy(template_path, file_path)
                if not silent:
                    MessageBoxWindow.message_simple(self, "Template Downloaded",f"Template saved to:\n{file_path}")
                return file_path
            return None

        except Exception as e:
            if not silent:
                MessageBoxWindow.message_simple(self, "Error", f"Failed to download template:\n{str(e)}")
            return None

    def download_md_tvd_template(self):
        """Download the MD-to-TVD template Excel file"""
        return self._download_template(
            "MD_TVD_Template.xlsx",
            "Save MD-to-TVD Template"
        )

    def download_interpretation_template(self, silent=False):
        """Download interpretation template, optionally return path without dialog"""
        return self._download_template(
            "Interpretation_Template.xlsx",
            "Save Interpretation Template",
            silent
        )

    def _get_template_path(self):
        """Helper for template path handling"""
        self.template_path = self.download_interpretation_template(silent=False)
        return bool(self.template_path)

    # Optimized data processing pipeline
    def process_data(self):
        """Streamlined data processing pipeline"""
        try:
            self._get_template_path()
            self.copy_statistics()
            self.paste_to_template(self.template_path)
            self.generate_as2_files()

            MessageBoxWindow.message_simple(self, "Processing Complete",
                "Data processed successfully!\n\n"
                f"Template saved as: {os.path.basename(self.template_path)}\n"
                "AS2 files generated for both gauges")
        except Exception as e:
            QMessageBox.critical(self, "Processing Error", f"Failed to process data:\n{str(e)}")

    # Optimized template pasting
    def paste_to_template(self, template_path):
        """Paste clipboard data to template starting at C13"""
        try:
            wb = openpyxl.load_workbook(template_path)
            sheet = wb.active
            clipboard = QApplication.clipboard()
            rows = clipboard.text().split('\n')

            for row_idx, row in enumerate(rows):
                if not row.strip():
                    continue

                cells = row.split('\t')[:14]  # Only process first 14 columns
                for col_idx, cell_value in enumerate(cells):
                    try:
                        value = float(cell_value)
                    except ValueError:
                        value = cell_value

                    sheet.cell(
                        row=13 + row_idx,
                        column=3 + col_idx,
                        value=value
                    )

            sheet.cell(row=3, column=3, value=f": {self.location}")
            sheet.cell(row=4, column=3, value=f": {self.well}")
            sheet.cell(row=4, column=18, value=self.sea_level)
            sheet.cell(row=5, column=18, value=self.bdf)

            wb.save(template_path)
            return True

        except Exception as e:
            QMessageBox.critical(
                self,
                "Paste Error",
                f"Failed to paste to template:\n{str(e)}"
            )
            return False

    def add_event(self):
        """Add a new event to the events list and table"""
        time_val = self.event_time_edit.time()
        time_str = time_val.toString("HH:mm:ss")
        desc = self.event_desc_edit.text().strip()

        if not desc:
            QMessageBox.warning(self, "Missing Description", "Please enter an event description")
            return

        # Get last time from data
        last_time = None
        if self.top_data:
            last_time = self.top_data[-1][0].time()
        elif self.bottom_data:
            last_time = self.bottom_data[-1][0].time()

        # Validate time
        if last_time and time_val > last_time:
            QMessageBox.warning(
                self,
                "Invalid Time",
                f"Event time exceeds last data point time ({last_time.toString('HH:mm:ss')})"
            )
            return

        # Add to events list
        self.events.append((time_str, desc))

        # Add to table
        row = self.event_table.rowCount()
        self.event_table.insertRow(row)
        self.event_table.setItem(row, 0, QTableWidgetItem(time_str))
        self.event_table.setItem(row, 1, QTableWidgetItem(desc))

        # Clear input fields
        self.event_desc_edit.clear()

        # Enable remove button
        self.remove_event_btn.setEnabled(True)

    def remove_event(self):
        """Remove selected event from the events list and table"""
        selected_row = self.event_table.currentRow()
        if selected_row >= 0:
            self.event_table.removeRow(selected_row)
            del self.events[selected_row]

            # Disable remove button if no events left
            if not self.events:
                self.remove_event_btn.setEnabled(False)

    # Optimized file dialog handling
    def open_file_dialog(self, file_type):
        """Unified file dialog opener"""
        file_types = {
            'tvd': ("Open TVD Calculation File", "Excel Files (*.xls *.xlsx)"),
            'top': ("Open Top Gauge Data File", "Text Files (*.txt)"),
            'bottom': ("Open Bottom Gauge Data File", "Text Files (*.txt)"),
            'timesheet': ("Open Survey Timesheet", "Excel Files (*.xls *.xlsx)")
        }

        title, file_filter = file_types.get(file_type, ("Open File", "All Files (*)"))
        file_path, _ = QFileDialog.getOpenFileName(self, title, "", file_filter)

        if file_path and hasattr(self.drag_drop_widget, 'set_file'):
            self.drag_drop_widget.set_file(file_type, file_path)
        return file_path

    def save_file(self):
        MessageBoxWindow.message_simple(self, "Save", "Save functionality will be implemented here")

    def init_zoom_toolbar(self):
        """Initialize zoom toolbar with reset button"""
        # Create toolbar widget
        self.toolbar_widget = QWidget()
        toolbar_layout = QHBoxLayout(self.toolbar_widget)
        toolbar_layout.setContentsMargins(0, 5, 0, 5)

        # Reset zoom button
        self.reset_zoom_btn = QPushButton("Reset Zoom")
        self.reset_zoom_btn.setStyleSheet("""
            QPushButton {
                background-color: red;
                border: 1px solid #aaa;
                border-radius: 5px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: pink;
            }
        """)
        self.reset_zoom_btn.setFixedHeight(30)
        self.reset_zoom_btn.clicked.connect(self.reset_graph_zoom)
        self.reset_zoom_btn.setEnabled(False)

        # Add copy graphs button next to reset zoom
        self.copy_graphs_button = QPushButton("Copy Graphs")
        self.copy_graphs_button.setIcon(QIcon(get_icon_path('copy')))
        self.copy_graphs_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.copy_graphs_button.clicked.connect(self.copy_graphs)
        self.copy_graphs_button.setEnabled(False)  # Disabled until data is loaded

        # Add to layout
        toolbar_layout.addWidget(self.reset_zoom_btn)
        toolbar_layout.addWidget(self.copy_graphs_button)
        # toolbar_layout.addStretch()

    # Add this method to SGSTXTApp
    def reset_graph_zoom(self):
        """Reset graph zoom to original view"""
        if self.initial_x_lim:
            ax_top = self.current_canvas.figure.axes[0]
            ax_bottom = self.current_canvas.figure.axes[1]

            ax_top.set_xlim(self.initial_x_lim)
            ax_bottom.set_xlim(self.initial_x_lim)
            self.current_canvas.draw_idle()
            self.reset_zoom_btn.setEnabled(False)

    # Add this method to SGSTXTApp
    def on_horizontal_select(self, xmin, xmax):
        """Handle horizontal zoom selection"""
        ax_top = self.current_canvas.figure.axes[0]
        ax_bottom = self.current_canvas.figure.axes[1]

        ax_top.set_xlim(xmin, xmax)
        ax_bottom.set_xlim(xmin, xmax)
        self.current_canvas.draw_idle()
        self.reset_zoom_btn.setEnabled(True)

    def process_all_files(self, top_file_path, bottom_file_path, timesheet_file_path, tvd_file_path):
        try:
            self.top_file_path = top_file_path
            self.bottom_file_path = bottom_file_path
            self.timesheet_file_path = timesheet_file_path
            self.tvd_file_path = tvd_file_path

            # Process TVD file first
            self.tvd_data = self.process_tvd_file(tvd_file_path)

            # Process timesheet to get station timings
            self.station_timings = self.process_excel_timesheet(timesheet_file_path)

            # Process both gauge data files
            self.top_data = self.process_sgs_txt_file(top_file_path)
            self.bottom_data = self.process_sgs_txt_file(bottom_file_path)

            # Reset events when processing new files
            self.events = []
            self.event_table.setRowCount(0)
            self.remove_event_btn.setEnabled(False)

            if self.top_data and self.bottom_data and self.station_timings and self.tvd_data:
                # Show results screen
                self.content_stack.setCurrentIndex(1)

                # Create stacked graphs
                self.show_stacked_graphs(top_file_path, bottom_file_path)

                # Populate table with station timings and gauge data
                self.populate_station_table()
            else:
                QMessageBox.critical(self, "Error", "Failed to process one or more files")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to process files:\n{str(e)}")

    def process_tvd_file(self, file_path):
        """Optimized TVD file processing with better validation"""
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb[wb.sheetnames[0]]
            elif lower_path.endswith('.xls'):
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
            else:
                QMessageBox.critical(self, "Error", "Unsupported TVD file format")
                return {}

            tvd_data = {
                'ahd_values': [],
                'tvd_values': []
            }
            self.bdf = self.get_cell_value(sheet, 1, 3)
            self.sea_level = self.get_cell_value(sheet, 2, 3)
            self.location = self.get_cell_value(sheet, 2, 8)
            self.well = self.get_cell_value(sheet, 3, 8)

            # Process rows more efficiently
            for row in range(8, sheet.max_row if hasattr(sheet, 'max_row') else sheet.nrows):
                ahd_val = self.parse_numeric_cell(self.get_cell_value(sheet, row, 2))
                tvd_val = self.parse_numeric_cell(self.get_cell_value(sheet, row, 8))

                if ahd_val is None and tvd_val is None:
                    break

                if ahd_val is not None:
                    tvd_data['ahd_values'].append(ahd_val)
                if tvd_val is not None:
                    tvd_data['tvd_values'].append(tvd_val)

            return tvd_data

        except Exception as e:
            QMessageBox.critical(self, "Error", f"TVD Processing Error:\n{str(e)}")
            return {}

    def get_cell_value(self, sheet, row, col):
        """Unified cell value getter for different Excel formats"""
        try:
            if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                return sheet.cell(row=row+1, column=col+1).value
            elif isinstance(sheet, xlrd.sheet.Sheet):
                return sheet.cell_value(row, col)
        except (IndexError, AttributeError):
            return None

    def parse_numeric_cell(self, value):
        """Parse numeric cell value with error handling"""
        if value is None:
            return None
        try:
            if isinstance(value, str):
                value = value.replace(',', '')
            return float(value)
        except (ValueError, TypeError):
            return None

    def generate_as2_files(self):
        """Generate AS2 files for both top and bottom gauges"""
        if not self.top_data or not self.bottom_data:
            QMessageBox.warning(self, "No Data", "No gauge data available to generate AS2 files")
            return

        try:
            top_as2_path = self.generate_as2_file(self.top_file_path, self.top_data)
            bottom_as2_path = self.generate_as2_file(self.bottom_file_path, self.bottom_data)

            if top_as2_path and bottom_as2_path:
                QMessageBox.information(
                    self,
                    "Files Created",
                    f"Successfully created AS2 files:\n\n"
                    f"Top Gauge: {top_as2_path.split('/')[-1]}\n"
                    f"Bottom Gauge: {bottom_as2_path.split('/')[-1]}",
                    QMessageBox.StandardButton.Ok
                )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create AS2 files:\n{str(e)}")

    def generate_as2_file(self, input_file_path, data):
        """Generate AS2 formatted file from processed data with right-aligned columns"""
        # Determine output file path
        if input_file_path.endswith('.txt'):
            output_file_path = input_file_path.replace('.txt', '.AS2')
        else:
            output_file_path = input_file_path + '.AS2'

        if not data:
            # Create empty file if no data
            with open(output_file_path, 'w'):
                pass
            return output_file_path

        # Create event lookup dictionary
        event_dict = {time: desc for time, desc in self.events}

        # Calculate column widths
        max_pressure_width = max(len(f"{p:.2f}") for _, p, _ in data) if data else 0
        max_temp_width = max(len(f"{t:.3f}") for _, _, t in data) if data else 0

        with open(output_file_path, 'w') as f:
            for dt, pressure, temperature in data:
                date_str = dt.strftime("%d/%m/%Y")
                time_str = dt.strftime("%H:%M:%S")

                # Format numbers with fixed decimals and right-align
                temp_str = f"{temperature:.2f}".rjust(max_temp_width)
                pressure_str = f"{pressure:.3f}".rjust(max_pressure_width)

                # Check if there's an event at this time
                event_desc = event_dict.get(time_str, "")

                # Format line with event if exists
                line = f"{date_str}  {time_str}    {temp_str}     {pressure_str}  {event_desc}\n"
                f.write(line)

        return output_file_path

    def on_mouse_move(self, event):
        """Handle mouse movement over the graph"""
        if event.inaxes is None:
            # Hide everything when mouse leaves plot area
            self.cursor_vline_top.set_visible(False)
            self.cursor_vline_bottom.set_visible(False)
            self.top_p_marker.set_visible(False)
            self.top_t_marker.set_visible(False)
            self.bottom_p_marker.set_visible(False)
            self.bottom_t_marker.set_visible(False)
            self.status_text.set_text("")
            self.cursor_vline_top.figure.canvas.draw_idle()
            return

        x = event.xdata
        x_dt = mdates.num2date(x)

        # Update vertical lines
        self.cursor_vline_top.set_xdata([x])
        self.cursor_vline_bottom.set_xdata([x])
        self.cursor_vline_top.set_visible(True)
        self.cursor_vline_bottom.set_visible(True)

        # Initialize status text
        status_lines = [f"Time: {x_dt.strftime('%H:%M:%S')}"]

        # Find and display top gauge values
        if self.top_times is not None:
            idx = np.argmin(np.abs(self.top_times - x))
            closest_time = self.top_times[idx]
            time_diff = abs(closest_time - x)

            # Only show if we're close to actual data point
            if time_diff < 0.0001:  # Matplotlib time units (days)
                top_p = self.top_pressures[idx]
                top_t = self.top_temps[idx]

                # Update markers
                self.top_p_marker.set_data([x], [top_p])
                self.top_t_marker.set_data([x], [top_t])
                self.top_p_marker.set_visible(True)
                self.top_t_marker.set_visible(True)

                status_lines.append(f"Top: Pressure = {top_p:.2f} psia, Temp = {top_t:.2f} °F")
            else:
                self.top_p_marker.set_visible(False)
                self.top_t_marker.set_visible(False)

        # Find and display bottom gauge values
        if self.bottom_times is not None:
            idx = np.argmin(np.abs(self.bottom_times - x))
            closest_time = self.bottom_times[idx]
            time_diff = abs(closest_time - x)

            # Only show if we're close to actual data point
            if time_diff < 0.0001:  # Matplotlib time units (days)
                bottom_p = self.bottom_pressures[idx]
                bottom_t = self.bottom_temps[idx]

                # Update markers
                self.bottom_p_marker.set_data([x], [bottom_p])
                self.bottom_t_marker.set_data([x], [bottom_t])
                self.bottom_p_marker.set_visible(True)
                self.bottom_t_marker.set_visible(True)

                status_lines.append(f"Bottom: Pressure = {bottom_p:.2f} psia, Temp = {bottom_t:.2f} °F")
            else:
                self.bottom_p_marker.set_visible(False)
                self.bottom_t_marker.set_visible(False)

        # Update status text
        self.status_text.set_text("\n".join(status_lines))

        # Redraw canvas
        self.cursor_vline_top.figure.canvas.draw_idle()
