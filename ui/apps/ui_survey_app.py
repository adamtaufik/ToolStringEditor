import datetime
import gc
import io
import os
import pickle
import re
import shutil
import tempfile

import pyqtgraph as pg
import openpyxl
import pandas as pd
import numpy as np
import xlrd
from PyQt6.QtCore import Qt, QTimer, QSize, QObject, QEvent, QDate, QParallelAnimationGroup, QPropertyAnimation, \
    QEasingCurve, QPoint
from PyQt6.QtGui import QDragEnterEvent, QDropEvent, QImage, QIcon, QColor
from PyQt6.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, QHBoxLayout, QApplication,
    QSplitter, QGridLayout, QGroupBox, QStackedWidget, QTimeEdit, QLineEdit, QToolButton, QFrame, QDateEdit,
    QGraphicsOpacityEffect
)
pg.setConfigOptions(background='w', antialias=True)

# Local imports
from ui.components.ui_footer import FooterWidget
from ui.components.ui_sidebar_widget import SidebarWidget
from ui.components.ui_titlebar import CustomTitleBar
from ui.windows.ui_messagebox_window import MessageBoxWindow
from utils.path_finder import get_icon_path, get_path
from utils.styles import GROUPBOX_STYLE, MODERN_GROUPBOX_STYLE, TEMPLATE_BUTTON, ACTION_BUTTON, DELETE_BUTTON
from utils.theme_manager import apply_theme, toggle_theme


class QuadrupleDragDropWidget(QWidget):
    def __init__(self, callback, parent=None):
        super().__init__(parent)
        self.callback = callback
        self.setAcceptDrops(True)

        # Define file type configurations
        self.file_types = {
            "top": {
                "label": "Drag & Drop Top Gauge Data File (.txt)",
                "drop_text": "Top Gauge Data\n(.txt)",
                "extensions": ('.txt',),
                "file_path": None
            },
            "bottom": {
                "label": "Drag & Drop Bottom Gauge Data File (.txt)",
                "drop_text": "Bottom Gauge Data\n(.txt)",
                "extensions": ('.txt',),
                "file_path": None
            },
            "timesheet": {
                "label": "Drag & Drop Survey Timesheet (.xls/.xlsx)",
                "drop_text": "Survey Timesheet\n(.xls/.xlsx)",
                "extensions": ('.xls', '.xlsx', '.XLS'),
                "file_path": None
            },
            "tvd": {
                "label": "Drag & Drop TVD Calculation File (.xls/.xlsx)",
                "drop_text": "TVD Calculation\n(.xls/.xlsx)",
                "extensions": ('.xls', '.xlsx', '.XLS'),
                "file_path": None
            }
        }

        layout = QGridLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)

        # Create drop areas and store references
        self.drop_areas = {}
        self.labels = {}
        self.clear_buttons = {}
        self.browse_buttons = {}

        for i, (file_type, config) in enumerate(self.file_types.items()):
            row = i // 2
            col = i % 2

            # Create UI components
            label = QLabel(config["label"])
            drop_area = self.create_drop_area(config["drop_text"])
            buttons_layout = self.create_file_buttons(file_type)

            # Create group box
            title = file_type.upper() + (" GAUGE" if file_type in ["top", "bottom"] else "")
            group_box = self.create_group_box(title, drop_area, label, buttons_layout)

            # Add to layout and store references
            layout.addWidget(group_box, row, col)
            self.drop_areas[file_type] = drop_area
            self.labels[file_type] = label

        # Process and Clear buttons
        button_container = QWidget()
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)

        self.clear_all_btn = self.create_button("Clear All", "#e74c3c", "#c0392b", 40)
        self.clear_all_btn.clicked.connect(self.clear_all_files)
        button_layout.addWidget(self.clear_all_btn)

        self.process_btn = self.create_button("Process Files", "#3498db", None, 40)
        self.process_btn.setEnabled(False)
        self.process_btn.clicked.connect(self.process_files)
        button_layout.addWidget(self.process_btn)

        layout.addWidget(button_container, 2, 0, 1, 2)

        # Initially hide all group boxes and bottom buttons for startup animation
        for group_box in self.findChildren(QGroupBox):
            group_box.setGraphicsEffect(QGraphicsOpacityEffect(group_box))
            group_box.graphicsEffect().setOpacity(0)
        button_container.setGraphicsEffect(QGraphicsOpacityEffect(button_container))
        button_container.graphicsEffect().setOpacity(0)

        # Animation trigger flag
        self.animation_played = False

        # --- Startup Animations ---
        QTimer.singleShot(200, self.trigger_startup_animation_once)

    def trigger_startup_animation_once(self):
        """Ensure animations only play once when the widget first becomes visible"""
        if not self.animation_played:
            self.animate_startup()
            self.animation_played = True

    # --- Animation Helpers ---#
    def animate_startup(self):
        """Sequentially animate all main widgets on startup"""
        delay = 0
        for group_box in self.findChildren(QGroupBox):
            QTimer.singleShot(delay, lambda gb=group_box: self.fade_and_slide_in(gb))
            delay += 150  # stagger each box slightly

        # Animate bottom button row after all boxes
        QTimer.singleShot(delay + 200, lambda: self.fade_widget(self.process_btn.parentWidget()))

    def fade_and_slide_in(self, widget):
        """Fade and slide-in animation for a widget"""
        start_pos = widget.pos() + QPoint(0, 30)
        end_pos = widget.pos()
        widget.move(start_pos)

        # Fade
        opacity = QGraphicsOpacityEffect(widget)
        widget.setGraphicsEffect(opacity)
        fade_anim = QPropertyAnimation(opacity, b"opacity", self)
        fade_anim.setDuration(400)
        fade_anim.setStartValue(0)
        fade_anim.setEndValue(1)
        fade_anim.setEasingCurve(QEasingCurve.Type.OutCubic)

        # Slide
        move_anim = QPropertyAnimation(widget, b"pos", self)
        move_anim.setDuration(400)
        move_anim.setStartValue(start_pos)
        move_anim.setEndValue(end_pos)
        move_anim.setEasingCurve(QEasingCurve.Type.OutCubic)

        # Combine
        group = QParallelAnimationGroup(self)
        group.addAnimation(fade_anim)
        group.addAnimation(move_anim)
        group.start(QParallelAnimationGroup.DeletionPolicy.DeleteWhenStopped)

        widget._anim_group = group  # prevent garbage collection

    def fade_widget(self, widget):
        """Simple fade-in animation for any widget"""
        opacity = QGraphicsOpacityEffect(widget)
        widget.setGraphicsEffect(opacity)
        anim = QPropertyAnimation(opacity, b"opacity", self)
        anim.setDuration(500)
        anim.setStartValue(0)
        anim.setEndValue(1)
        anim.setEasingCurve(QEasingCurve.Type.OutQuad)
        anim.start(QPropertyAnimation.DeletionPolicy.DeleteWhenStopped)
        widget._fade_anim = anim


    # --- Helper Methods ---#
    def create_button(self, text, color, hover_color=None, height=None):
        """Create standardized buttons with consistent styling"""
        button = QPushButton(text)
        if height:
            button.setMinimumHeight(height)

        style = f"""
            QPushButton {{
                background-color: {color};
                color: white;
                font-weight: bold;
                border-radius: 10px;
                padding: 8px;
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
                color: #888888;
            }}
        """

        if hover_color:
            style += f"""
            QPushButton:hover {{
                background-color: {hover_color};
            }}
            """

        button.setStyleSheet(style)
        return button

    def create_drop_area(self, text):
        """Create a consistent drop area widget"""
        drop_area = QLabel(text)
        drop_area.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px dashed #aaa;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(240, 240, 240, 100);
            }
        """)
        drop_area.setMinimumSize(200, 180)
        drop_area.setAcceptDrops(True)
        drop_area.dragEnterEvent = self.dragEnterEvent
        drop_area.dropEvent = lambda event, da=drop_area: self.dropEvent(event, da)
        return drop_area

    def create_file_buttons(self, file_type):
        """Create browse/clear buttons for a file type"""
        button_layout = QHBoxLayout()

        browse_btn = self.create_button("Browse", "#3498db", "#2980b9", 25)
        browse_btn.setStyleSheet(browse_btn.styleSheet() + "color: black;")
        browse_btn.clicked.connect(lambda _, ft=file_type: self.browse_file(ft))

        clear_btn = self.create_button("Clear", "#f8d7da", "#f5c6cb", 25)
        clear_btn.setStyleSheet(clear_btn.styleSheet() + "color: #721c24;")
        clear_btn.setEnabled(False)
        clear_btn.clicked.connect(lambda _, ft=file_type: self.clear_file(ft))

        button_layout.addWidget(browse_btn)
        button_layout.addWidget(clear_btn)

        # Store button references
        self.browse_buttons[file_type] = browse_btn
        self.clear_buttons[file_type] = clear_btn

        return button_layout

    def create_group_box(self, title, widget, label, buttons):
        """Create a standardized group box container"""
        group = QGroupBox(title)
        group_layout = QVBoxLayout(group)
        group_layout.addWidget(label)
        group_layout.addWidget(widget)
        group_layout.addLayout(buttons)
        group.setStyleSheet(GROUPBOX_STYLE)
        return group

    # --- File Handling ---#
    def handle_file_drop(self, drop_area, file_path):
        """Handle file drop on a specific area"""
        for file_type, config in self.file_types.items():
            if drop_area == self.drop_areas[file_type]:
                if file_path.lower().endswith(config["extensions"]):
                    self.set_file(file_type, file_path)
                    return

        MessageBoxWindow.message_simple(self, "Invalid File", "Please drop the correct file type in each area")

    def browse_file(self, file_type):
        """Open file dialog for specific file type"""
        config = self.file_types[file_type]
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Open {file_type.capitalize()} File",
            "",
            f"Files ({' '.join(['*' + ext for ext in config['extensions']])})"
        )
        if file_path:
            self.set_file(file_type, file_path)

    def set_file(self, file_type, file_path):
        """Set file for a specific type and update UI"""
        config = self.file_types[file_type]
        config["file_path"] = file_path
        filename = file_path.split('/')[-1]

        drop_area = self.drop_areas[file_type]
        drop_area.setText(f"Loaded:\n{filename}")
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px solid #2ecc71;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(200, 255, 200, 150);
            }
        """)

        self.clear_buttons[file_type].setEnabled(True)
        self.update_process_button()

    def clear_file(self, file_type):
        """Clear file for a specific type and reset UI"""
        config = self.file_types[file_type]
        config["file_path"] = None

        drop_area = self.drop_areas[file_type]
        drop_area.setText(config["drop_text"])
        drop_area.setStyleSheet("""
            QLabel {
                border: 3px dashed #aaa;
                border-radius: 15px;
                font: bold 12pt 'Segoe UI';
                padding: 30px;
                background-color: rgba(240, 240, 240, 100);
            }
        """)

        self.clear_buttons[file_type].setEnabled(False)
        self.update_process_button()

    def clear_all_files(self):
        """Clear all uploaded files and reset UI"""
        for file_type in self.file_types:
            self.clear_file(file_type)

    def update_process_button(self):
        """Enable process button only when all files are loaded"""
        all_loaded = all(
            config["file_path"] is not None
            for config in self.file_types.values()
        )
        self.process_btn.setEnabled(all_loaded)

    # --- Event Handlers ---#
    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
            widget = self.childAt(event.position().toPoint())
            if widget in self.drop_areas.values():
                widget.setStyleSheet("""
                    QLabel {
                        border: 3px dashed #3498db;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(200, 230, 255, 150);
                    }
                """)

    def dropEvent(self, event: QDropEvent, drop_area):
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_path = urls[0].toLocalFile()
            self.handle_file_drop(drop_area, file_path)
            # Reset style after drop
            if self.file_types[[k for k, v in self.drop_areas.items() if v == drop_area][0]]["file_path"]:
                drop_area.setStyleSheet("""
                    QLabel {
                        border: 3px solid #2ecc71;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(200, 255, 200, 150);
                    }
                """)
            else:
                drop_area.setStyleSheet("""
                    QLabel {
                        border: 3px dashed #aaa;
                        border-radius: 15px;
                        font: bold 12pt 'Segoe UI';
                        padding: 30px;
                        background-color: rgba(240, 240, 240, 100);
                    }
                """)

    def process_files(self):
        """Process files when all are loaded"""
        if all(config["file_path"] for config in self.file_types.values()):
            self.callback(
                self.file_types["top"]["file_path"],
                self.file_types["bottom"]["file_path"],
                self.file_types["timesheet"]["file_path"],
                self.file_types["tvd"]["file_path"]
            )

class HoverCursorFilter(QObject):
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.Enter:
            QApplication.setOverrideCursor(Qt.CursorShape.PointingHandCursor)
            return True
        elif event.type() == QEvent.Type.Leave:
            QApplication.restoreOverrideCursor()
            return True
        return super().eventFilter(obj, event)

class SGSTXTApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("SGS / FGS txt processing")
        self.setMinimumSize(1400, 800)
        self.save_file_path = None
        self.start_time = None
        self.date = None
        self.location = None
        self.well = None
        self.current_theme = "Deleum"
        self.sidebar_expanded = False
        self.station_timings = []
        self.top_data = []
        self.bottom_data = []
        self.top_file_path = None
        self.bottom_file_path = None
        self.timesheet_file_path = None
        self.ahd_tvd_map = {}
        self.events = []
        self.current_canvas = None
        self.bdf = 0
        self.sea_level = 0

        self.groupbox_styles = {
            "Deleum": {
                "text_color": "white",
                "accent_color": "#4FC3F7"
            },
            "Dark": {
                "text_color": "white",
                "accent_color": "#0D47A1"
            }
        }
        self.init_ui()

    def init_ui(self):
        main_container = QVBoxLayout(self)
        main_container.setContentsMargins(0, 0, 0, 0)
        main_container.setSpacing(0)

        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)

        items = [
            (get_icon_path('save'), "Save Survey", self.save_file, "Save the current survey (Ctrl+S)"),
            (get_icon_path('load'), "Load Survey", self.open_file, "Open a survey (Ctrl+O)"),
            (get_icon_path('export'), "Process && Export", self.process_data, "Export to Interpretation File")
        ]

        self.sidebar = SidebarWidget(self, items)
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint)
        self.title_bar = CustomTitleBar(self, lambda: self.sidebar.toggle_visibility(), "SGS/FGS txt processing")

        main_container.addWidget(self.title_bar)
        main_layout.addWidget(self.sidebar)

        # Main content area - using stacked widget for screens
        self.content_stack = QStackedWidget()
        main_layout.addWidget(self.content_stack)

        # File upload screen
        self.file_upload_widget = QWidget()
        file_upload_layout = QVBoxLayout(self.file_upload_widget)
        file_upload_layout.setContentsMargins(10, 10, 10, 10)

        upload_container_layout = QHBoxLayout()

        # Triple drag and drop area
        self.drag_drop_widget = QuadrupleDragDropWidget(self.process_all_files, self)
        upload_container_layout.addWidget(self.drag_drop_widget, 1)

        # Create templates section for file upload screen
        templates_container = QWidget()
        templates_layout = QVBoxLayout(templates_container)
        templates_layout.setContentsMargins(0, 0, 0, 0)

        # Group Box: Template Downloads
        template_group = QGroupBox("Download Blank Templates")
        template_layout = QVBoxLayout(template_group)
        template_layout.setSpacing(8)

        download_template_button = QPushButton("Interpretation")
        download_template_button.setIcon(QIcon(get_icon_path('download')))
        download_template_button.setStyleSheet(TEMPLATE_BUTTON)
        download_template_button.clicked.connect(self.download_interpretation_template)

        download_md_tvd_button = QPushButton("TVD Calculation")
        download_md_tvd_button.setIcon(QIcon(get_icon_path('download')))
        download_md_tvd_button.setStyleSheet(TEMPLATE_BUTTON)
        download_md_tvd_button.clicked.connect(self.download_md_tvd_template)

        download_timesheet_button = QPushButton("Timesheet")
        download_timesheet_button.setIcon(QIcon(get_icon_path('download')))
        download_timesheet_button.setStyleSheet(TEMPLATE_BUTTON)
        download_timesheet_button.clicked.connect(self.download_timesheet_template)

        template_layout.addWidget(download_template_button)
        template_layout.addWidget(download_md_tvd_button)
        template_layout.addWidget(download_timesheet_button)

        instruction_group = QGroupBox("Quick Guide")
        instruction_layout = QVBoxLayout(instruction_group)

        # Define your steps
        instruction_steps = [
            ("Step 1:", "Open PPS SmartView/MetroWin and convert raw (.rec/.p3w) files into .txt files"),
            ("Step 2:", "Download and fill in the TVD Calculation template"),
            ("Step 3:", "Drag and drop Top Gauge and Bottom Gauge .txt files"),
            ("Step 4:", "Drag and drop filled TVD Calculation file"),
            ("Step 5:", "Drag and drop Time Sheet"),
            ("Step 6:", "Click 'Process Files'")
        ]

        # Hanging indent CSS
        style = """
        <div style='margin-left: 0px; text-indent: -45px; margin-left: 45px; font-family: "Segoe UI", sans-serif; font-size: 14px; line-height: 1.5; color: white;'>
            <i>{step}</i> {desc}
        </div>
        """

        for step, desc in instruction_steps:
            label = QLabel()
            label.setTextFormat(Qt.TextFormat.RichText)
            label.setText(style.format(step=step, desc=desc))
            label.setWordWrap(True)
            instruction_layout.addWidget(label)

        # Add group box to layout
        # templates_layout.addStretch()
        templates_layout.addWidget(template_group)
        templates_layout.addWidget(instruction_group)
        templates_layout.addStretch()

        # Add templates section to file upload layout
        upload_container_layout.addWidget(templates_container)

        file_upload_layout.addLayout(upload_container_layout)

        # Apply theme styling
        current_style = self.groupbox_styles[self.current_theme]
        style = MODERN_GROUPBOX_STYLE.format(
            text_color=current_style["text_color"],
            accent_color=current_style["accent_color"]
        )
        template_group.setStyleSheet(style)
        instruction_group.setStyleSheet(style)

        # Footer
        self.footer = FooterWidget(self, theme_callback=self.toggle_theme)
        self.theme_button = self.footer.theme_button
        file_upload_layout.addWidget(self.footer)

        self.content_stack.addWidget(self.file_upload_widget)

        # Results display screen
        self.results_widget = QWidget()
        results_layout = QVBoxLayout(self.results_widget)
        results_layout.setContentsMargins(10, 10, 10, 10)

        # Navigation bar at top of results
        nav_bar = QWidget()
        nav_layout = QHBoxLayout(nav_bar)
        nav_layout.setContentsMargins(0, 0, 0, 10)

        self.back_button = QPushButton("← Back to File Upload")
        self.back_button.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        self.back_button.setFixedHeight(40)
        self.back_button.clicked.connect(self.show_file_upload)
        nav_layout.addWidget(self.back_button)

        results_layout.addWidget(nav_bar)

        # Results content
        results_content = QWidget()
        results_content_layout = QHBoxLayout(results_content)
        results_content_layout.setContentsMargins(0, 0, 0, 0)

        # Graph area - will contain two stacked graphs
        self.graph_widget = QWidget()
        self.graph_layout = QVBoxLayout(self.graph_widget)
        self.graph_layout.setContentsMargins(0, 0, 10, 0)

        # Table area - now in a group box
        table_container = QWidget()
        table_layout = QVBoxLayout(table_container)
        table_layout.setContentsMargins(0, 0, 0, 0)

        # Create group box for the table
        table_group = QGroupBox("Station Data")
        table_group_layout = QVBoxLayout(table_group)

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(18)
        self.table_widget.setHorizontalHeaderLabels([
            "Station", "Depth (ft)", "Start T.", "End T.",
            "AHD FTBDF", "TVD FTBDF",
            "High P (T)", "Low P (T)", "Med P (T)", "High T (T)", "Low T (T)", "Med T (T)",
            "High P (B)", "Low P (B)", "Med P (B)", "High T (B)", "Low T (B)", "Med T (B)"
        ])
        self.table_widget.horizontalHeader().setStretchLastSection(True)
        self.table_widget.setStyleSheet("""
            QTableWidget {
                background-color: #f8f8f8;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                padding: 4px;
                border: 1px solid #ccc;
                font-weight: bold;
            }
        """)
        self.table_widget.itemChanged.connect(self.on_station_time_changed)

        # Add table to group
        table_group_layout.addWidget(self.table_widget)

        info_group = QGroupBox("Survey Info")
        info_layout = QVBoxLayout(info_group)

        self.location_label = QLabel("Location      : ")
        self.well_label = QLabel("Well No.      : ")
        self.date_label = QLabel("Date of Survey: ")
        self.time_label = QLabel("Start Time    : ")
        self.bdf_label = QLabel("THF           : ")
        self.sea_level_label = QLabel("DFE     : ")

        info_layout.addWidget(self.location_label)
        info_layout.addWidget(self.well_label)
        info_layout.addWidget(self.date_label)
        info_layout.addWidget(self.time_label)
        info_layout.addWidget(self.bdf_label)
        info_layout.addWidget(self.sea_level_label)

        # Create button row with two group boxes
        button_row = QWidget()
        button_layout = QHBoxLayout(button_row)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(10)  # Add spacing between group boxes

        # Group Box 1: Template Downloads
        template_group = QGroupBox("Download Blank Templates")
        template_layout = QVBoxLayout(template_group)
        template_layout.setSpacing(8)  # Add spacing between buttons

        self.download_template_button = QPushButton("Interpretation")
        self.download_template_button.setIcon(QIcon(get_icon_path('download')))
        self.download_template_button.setStyleSheet(TEMPLATE_BUTTON)
        self.download_template_button.clicked.connect(self.download_interpretation_template)

        self.download_md_tvd_button = QPushButton("TVD Calculation")
        self.download_md_tvd_button.setIcon(QIcon(get_icon_path('download')))
        self.download_md_tvd_button.setStyleSheet(TEMPLATE_BUTTON)
        self.download_md_tvd_button.clicked.connect(self.download_md_tvd_template)

        template_layout.addWidget(self.download_template_button)
        template_layout.addWidget(self.download_md_tvd_button)

        # Group Box 2: Actions
        action_group = QGroupBox("Actions")
        action_layout = QVBoxLayout(action_group)
        action_layout.setSpacing(8)

        self.copy_button = QPushButton()
        self.copy_button.setIcon(QIcon(get_icon_path('copy')))  # Add copy icon
        self.copy_button.setText("Copy Statistics")
        self.copy_button.setStyleSheet(ACTION_BUTTON)
        self.copy_button.clicked.connect(self.copy_statistics)
        self.copy_button.setEnabled(False)  # Disabled until data is loaded

        self.generate_as2_button = QPushButton("Generate AS2 Files")
        self.generate_as2_button.setStyleSheet(ACTION_BUTTON)
        self.generate_as2_button.clicked.connect(self.generate_as2_files)
        self.generate_as2_button.setEnabled(False)  # Disabled until data is loaded

        self.process_data_button = QToolButton()
        self.process_data_button.setIcon(QIcon(get_icon_path('export')))
        self.process_data_button.setText("Process && Export")
        self.process_data_button.setToolButtonStyle(Qt.ToolButtonStyle.ToolButtonTextUnderIcon
)
        self.process_data_button.setIconSize(QSize(32, 32))  # Increase icon size
        self.process_data_button.setStyleSheet("""
            QToolButton {
                background-color: #28a745;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QToolButton:hover {
                background-color: #218838;
            }
            QToolButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.process_data_button.clicked.connect(self.process_data)
        self.process_data_button.setEnabled(False)  # Disabled until data is loaded

        action_layout.addWidget(self.copy_button)
        action_layout.addWidget(self.generate_as2_button)

        # Add group boxes to the button layout
        button_layout.addWidget(template_group, 1)
        button_layout.addWidget(action_group, 1)
        button_layout.addWidget(self.process_data_button)

        # Create event section
        event_group = QGroupBox("Events")
        event_layout = QGridLayout(event_group)

        # Event input fields
        self.event_time_edit = QTimeEdit()
        self.event_time_edit.setDisplayFormat("HH:mm:ss")

        # Event date input
        self.event_date_edit = QDateEdit()
        self.event_date_edit.setCalendarPopup(True)
        self.event_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.event_date_edit.setDate(QDate.currentDate())

        self.event_desc_edit = QLineEdit()
        self.event_desc_edit.setPlaceholderText("Enter event description")

        self.add_event_btn = QPushButton("Add")
        self.add_event_btn.setIcon(QIcon(get_icon_path('add')))
        self.add_event_btn.clicked.connect(self.add_event)
        self.add_event_btn.setStyleSheet("""
            QPushButton {
                background-color: #5cb85c;
                color: white;
                border-radius: 4px;
                padding: 5px;
            }
            QPushButton:hover {
                background-color: #4cae4c;
            }
        """)

        self.remove_event_btn = QPushButton("Remove")
        self.remove_event_btn.setIcon(QIcon(get_icon_path('remove')))
        self.remove_event_btn.clicked.connect(self.remove_event)
        self.remove_event_btn.setStyleSheet(DELETE_BUTTON)

        # Event table
        self.event_table = QTableWidget(0, 2)
        self.event_table.setHorizontalHeaderLabels(["Time", "Event Description"])
        self.event_table.setColumnWidth(0, 140)
        self.event_table.horizontalHeader().setStretchLastSection(True)
        self.event_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        # Add widgets to layout
        event_layout.addWidget(QLabel("Date:"), 0, 0)
        event_layout.addWidget(self.event_date_edit, 0, 1)

        event_layout.addWidget(QLabel("Time:"), 0, 2)
        event_layout.addWidget(self.event_time_edit, 0, 3)
        event_layout.addWidget(self.add_event_btn, 0, 5)

        event_layout.addWidget(QLabel("Description:"), 1, 0)
        event_layout.addWidget(self.event_desc_edit, 1, 1, 1, 4)
        event_layout.addWidget(self.remove_event_btn, 1, 5)
        event_layout.addWidget(self.event_table, 2, 0, 2, 6)

        # Add event section to table container
        table_layout.addWidget(event_group)
        table_layout.addWidget(button_row)

        # Inside the button_row section for results screen:
        self.copy_graphs_button = QPushButton("Copy Graphs")
        self.copy_graphs_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.copy_graphs_button.clicked.connect(self.copy_graphs)
        self.copy_graphs_button.setEnabled(False)  # Disabled until data is loaded

        # Add group to container
        table_layout.setSpacing(5)
        table_layout.setContentsMargins(10,0,0,0)
        info_event_row = QWidget()
        info_event_layout = QHBoxLayout(info_event_row)
        info_event_layout.setContentsMargins(0, 0, 0, 0)
        info_event_layout.setSpacing(10)

        info_event_layout.addWidget(info_group, 1)
        info_event_layout.addWidget(event_group, 2)

        table_layout.addWidget(info_event_row)
        table_layout.addWidget(table_group)  # Now using group box
        table_layout.addWidget(button_row)

        # Add widgets to splitter
        splitter = QSplitter(Qt.Orientation.Horizontal)
        splitter.addWidget(self.graph_widget)
        splitter.addWidget(table_container)
        splitter.setSizes([500, 500])
        results_content_layout.addWidget(splitter)

        results_layout.addWidget(results_content, 1)

        # Footer for results screen
        results_footer = FooterWidget(self, theme_callback=self.toggle_theme)
        results_layout.addWidget(results_footer)

        self.content_stack.addWidget(self.results_widget)

        main_container.addLayout(main_layout)
        self.setLayout(main_container)

        current_style = self.groupbox_styles[self.current_theme]
        print(current_style)
        style = MODERN_GROUPBOX_STYLE.format(
            text_color=current_style["text_color"],
            accent_color=current_style["accent_color"]
        )

        info_group.setStyleSheet(style)
        table_group.setStyleSheet(style)
        template_group.setStyleSheet(style)
        action_group.setStyleSheet(style)
        event_group.setStyleSheet(style)

        # Create toolbar widget for graphs
        self.toolbar_widget = QWidget()
        toolbar_layout = QHBoxLayout(self.toolbar_widget)
        toolbar_layout.setContentsMargins(0, 5, 0, 5)

        # Reset zoom button
        self.reset_zoom_btn = QPushButton("Reset")
        self.reset_zoom_btn.setIcon(QIcon(get_icon_path('unzoom')))
        self.reset_zoom_btn.setStyleSheet(DELETE_BUTTON)
        self.reset_zoom_btn.clicked.connect(self.reset_graph_zoom)
        self.reset_zoom_btn.setEnabled(False)

        # Copy graphs button
        self.copy_graphs_button = QPushButton("Copy")
        self.copy_graphs_button.setIcon(QIcon(get_icon_path('copy')))
        self.copy_graphs_button.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                border-radius: 4px;
                padding: 5px 10px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #8e44ad;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        """)
        self.copy_graphs_button.clicked.connect(self.copy_graphs)
        self.copy_graphs_button.setEnabled(False)

        # Add to layout
        toolbar_layout.addWidget(self.reset_zoom_btn)
        toolbar_layout.addWidget(self.copy_graphs_button)

        # Apply theme after UI is set up
        apply_theme(self, self.current_theme)

        # Create cursor filter
        self.cursor_filter = HoverCursorFilter()

        # Apply to all buttons
        for widget in self.findChildren(QPushButton) + self.findChildren(QToolButton):
            widget.installEventFilter(self.cursor_filter)

    def show_stacked_graphs(self, top_file_path, bottom_file_path):
        # Clear previous graph
        while self.graph_layout.count():
            item = self.graph_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # ============================
        # Main container: HBox → graphs | right column (legend + buttons)
        # ============================
        graph_container = QWidget()
        main_layout = QHBoxLayout(graph_container)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(10)

        # ----------------------------
        # Left side: stacked graphs
        # ----------------------------
        graphs_container = QWidget()
        graphs_layout = QVBoxLayout(graphs_container)
        graphs_layout.setContentsMargins(0, 0, 0, 0)
        graphs_layout.setSpacing(5)

        # ============================
        # Data extraction and conversion
        # ============================
        # Initialize data arrays
        top_timestamps = top_pressures = top_temps = None
        bottom_timestamps = bottom_pressures = bottom_temps = None

        # Process top data
        if self.top_data is not None:
            if hasattr(self.top_data, 'dtype') and self.top_data.dtype.names is not None:
                top_timestamps = np.array([dt.timestamp() for dt in self.top_data['datetime']])
                top_pressures = self.top_data['pressure'].astype(np.float32)
                top_temps = self.top_data['temperature'].astype(np.float32)
            elif isinstance(self.top_data, list) and len(self.top_data) > 0:
                top_timestamps = np.array([dt.timestamp() for dt, _, _ in self.top_data])
                top_pressures = np.array([p for _, p, _ in self.top_data], dtype=np.float32)
                top_temps = np.array([t for _, _, t in self.top_data], dtype=np.float32)

        # Process bottom data
        if self.bottom_data is not None:
            if hasattr(self.bottom_data, 'dtype') and self.bottom_data.dtype.names is not None:
                bottom_timestamps = np.array([dt.timestamp() for dt in self.bottom_data['datetime']])
                bottom_pressures = self.bottom_data['pressure'].astype(np.float32)
                bottom_temps = self.bottom_data['temperature'].astype(np.float32)
            elif isinstance(self.bottom_data, list) and len(self.bottom_data) > 0:
                bottom_timestamps = np.array([dt.timestamp() for dt, _, _ in self.bottom_data])
                bottom_pressures = np.array([p for _, p, _ in self.bottom_data], dtype=np.float32)
                bottom_temps = np.array([t for _, _, t in self.bottom_data], dtype=np.float32)

        # Store for later use
        self.top_timestamps = top_timestamps
        self.top_pressures = top_pressures
        self.top_temps = top_temps
        self.bottom_timestamps = bottom_timestamps
        self.bottom_pressures = bottom_pressures
        self.bottom_temps = bottom_temps

        # ============================
        # Custom Time Axis Formatter
        # ============================
        class TimeAxisItem(pg.AxisItem):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.enableAutoSIPrefix(False)

            def tickStrings(self, values, scale, spacing):
                strings = []
                for v in values:
                    try:
                        dt = datetime.datetime.fromtimestamp(v)
                        strings.append(dt.strftime("%H:%M:%S"))
                    except:
                        strings.append(str(v))
                return strings

        # ============================
        # Helper to create plot with secondary y-axis
        # ============================
        def create_plot_with_temp(timestamps, pressures, temps, title):
            # Enable OpenGL for faster rendering
            plot_widget = pg.PlotWidget(axisItems={'bottom': TimeAxisItem(orientation='bottom')}, useOpenGL=True)
            plot_widget.setLabel('left', 'Pressure', units='psia')
            plot_widget.setLabel('bottom', 'Time')
            plot_widget.setTitle(title)
            plot_widget.showGrid(x=True, y=True, alpha=0.3)

            # Add legend - THIS IS THE KEY ADDITION
            plot_widget.addLegend()

            # Main pressure curve - use OpenGL for large datasets
            pressure_curve = pg.PlotCurveItem(timestamps, pressures,
                                              pen=pg.mkPen('b', width=1.5),
                                              name='Pressure',  # Name appears in legend
                                              connect='all',
                                              antialias=True,
                                              useOpenGL=True)
            plot_widget.addItem(pressure_curve)

            # Secondary y-axis for temperature
            temp_view = pg.ViewBox()
            plot_widget.getPlotItem().scene().addItem(temp_view)
            plot_widget.getPlotItem().getAxis('right').linkToView(temp_view)
            temp_view.setXLink(plot_widget.getPlotItem())

            temp_curve = pg.PlotCurveItem(timestamps, temps,
                                          pen=pg.mkPen('r', width=1.5),
                                          name='Temperature',  # Name appears in legend
                                          connect='all',
                                          antialias=True,
                                          useOpenGL=True)
            temp_view.addItem(temp_curve)

            # IMPORTANT: Also add temperature curve to main plot's legend
            # This ensures it shows up in the legend even though it's in a different ViewBox
            plot_widget.getPlotItem().legend.addItem(temp_curve, temp_curve.name())

            # Keep views synced on resize
            def update_views():
                temp_view.setGeometry(plot_widget.getPlotItem().vb.sceneBoundingRect())
                temp_view.linkedViewChanged(plot_widget.getPlotItem().vb, temp_view.XAxis)

            plot_widget.getPlotItem().vb.sigResized.connect(update_views)

            # Right axis label
            temp_axis = pg.AxisItem('right')
            plot_widget.getPlotItem().layout.addItem(temp_axis, 2, 3)
            temp_axis.linkToView(temp_view)
            temp_axis.setLabel('Temperature', units='°F', color='r')

            return plot_widget

        # Add top and bottom plots to graphs_layout
        if top_timestamps is not None and len(top_timestamps) > 0:
            self.top_plot = create_plot_with_temp(top_timestamps, top_pressures, top_temps,
                                                  f"Top Gauge - {os.path.basename(top_file_path)}")
            graphs_layout.addWidget(self.top_plot)

        if bottom_timestamps is not None and len(bottom_timestamps) > 0:
            self.bottom_plot = create_plot_with_temp(bottom_timestamps, bottom_pressures, bottom_temps,
                                                     f"Bottom Gauge - {os.path.basename(bottom_file_path)}")
            if hasattr(self, 'top_plot') and self.top_plot is not None:
                self.bottom_plot.setXLink(self.top_plot)
            graphs_layout.addWidget(self.bottom_plot)

        main_layout.addWidget(graphs_container, stretch=3)

        # ============================
        # Station shading
        # ============================
        self.plot_regions = []

        if hasattr(self, 'station_timings') and self.station_timings:
            colors = []
            station_labels = []
            cmap = pg.colormap.get('viridis', source='matplotlib')  # can pick another colormap
            n = len(self.station_timings)
            for i in range(len(self.station_timings)):
                hue = int(360 * i / n)  # evenly spaced hues
                color = QColor()
                color.setHsv(hue, 200, 255, 60)
                # color.setAlpha(50)  # adjust 0-255 for transparency

                colors.append(color)

                # Get station label
                label = self.station_timings[i].get('label', f'Station {i+1}')
                station_labels.append(label)

            for idx, station in enumerate(self.station_timings):
                try:
                    start_ts = station['start'].timestamp()
                    end_ts = station['end'].timestamp()
                    color = colors[idx]
                    region_top = pg.LinearRegionItem(values=[start_ts, end_ts], brush=pg.mkBrush(color), movable=False)
                    region_bottom = pg.LinearRegionItem(values=[start_ts, end_ts], brush=pg.mkBrush(color),
                                                        movable=False)

                    if hasattr(self, 'top_plot') and self.top_plot is not None:
                        self.top_plot.addItem(region_top)
                    if hasattr(self, 'bottom_plot') and self.bottom_plot is not None:
                        self.bottom_plot.addItem(region_bottom)

                    self.plot_regions.append((region_top, region_bottom))
                except Exception as e:
                    print(f"Error adding station region {idx}: {e}")

        # ----------------------------
        # Right side: legend + buttons
        # ----------------------------
        right_column = QWidget()
        right_column.setFixedWidth(100)
        right_layout = QVBoxLayout(right_column)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        # Right side: legend + buttons
        if hasattr(self, 'station_timings') and self.station_timings:
            colors = []
            labels = []

            cmap = pg.colormap.get('viridis', source='matplotlib')  # can pick another colormap
            n = len(self.station_timings)
            for i, station in enumerate(self.station_timings):
                hue = int(360 * i / n)  # evenly spaced hues
                color = QColor()
                color.setHsv(hue, 200, 255, 60)
                # color.setAlpha(50)  # adjust 0-255 for transparency

                colors.append(color)

                # Include depth in label like old code
                station_name = station.get('station', f'Station {i + 1}')
                depth = station.get('depth', '')
                if depth != '':
                    label = f"{station_name} - {depth}ft"
                else:
                    label = station_name
                labels.append(label)

            self.create_station_legend(colors, labels, right_column)

        right_layout.addStretch()
        # ============================
        # Cursor readout widget
        # ============================
        self.cursor_readout_widget = QWidget()
        cursor_layout = QVBoxLayout(self.cursor_readout_widget)
        cursor_layout.setContentsMargins(0, 0, 0, 0)
        cursor_layout.setSpacing(4)

        self.cursor_time_label = QLabel("Time: --:--:--\n")
        self.cursor_top_label = QLabel("Top: \nP = -- psia\nT = -- °F\n")
        self.cursor_bottom_label = QLabel("Bottom: \nP = -- psia\nT = -- °F")

        for lbl in (
                self.cursor_time_label,
                self.cursor_top_label,
                self.cursor_bottom_label
        ):
            lbl.setStyleSheet(
                "font-family: Consolas; font-size: 11px;"
            )

        title = QLabel("Cursor Values")
        title.setStyleSheet("font-weight: bold;")

        cursor_layout.addWidget(title)
        cursor_layout.addWidget(self.cursor_time_label)
        cursor_layout.addWidget(self.cursor_top_label)
        cursor_layout.addWidget(self.cursor_bottom_label)

        right_layout.addWidget(self.cursor_readout_widget)

        # Toolbar buttons
        button_container = QWidget()
        button_layout = QVBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(5)

        # Reset Zoom button
        self.reset_zoom_btn.setEnabled(hasattr(self, 'top_plot') and self.top_plot is not None)
        button_layout.addWidget(self.reset_zoom_btn)

        # Copy Graphs button
        self.copy_graphs_button.setEnabled(True)
        button_layout.addWidget(self.copy_graphs_button)

        right_layout.addWidget(button_container)


        main_layout.addWidget(right_column, stretch=1)  # less width than graphs

        self.graph_layout.addWidget(graph_container)

        # ============================
        # Synchronized cursors
        # ============================
        if hasattr(self, 'top_plot') and self.top_plot is not None and \
                hasattr(self, 'bottom_plot') and self.bottom_plot is not None:

            # Create vertical lines for both plots
            self.cursor_vline_top = pg.InfiniteLine(angle=90, movable=False,
                                                    pen=pg.mkPen('k', width=1, style=pg.QtCore.Qt.PenStyle.DashLine))
            self.cursor_vline_bottom = pg.InfiniteLine(angle=90, movable=False,
                                                       pen=pg.mkPen('k', width=1, style=pg.QtCore.Qt.PenStyle.DashLine))

            self.top_plot.addItem(self.cursor_vline_top)
            self.bottom_plot.addItem(self.cursor_vline_bottom)

            # One handler, two scenes
            self.proxy_top = pg.SignalProxy(
                self.top_plot.scene().sigMouseMoved,
                rateLimit=60,
                slot=self.on_mouse_move_pyqtgraph
            )

            self.proxy_bottom = pg.SignalProxy(
                self.bottom_plot.scene().sigMouseMoved,
                rateLimit=60,
                slot=self.on_mouse_move_pyqtgraph
            )

    def on_mouse_move_pyqtgraph(self, event):
        """Handle mouse movement over either top or bottom PyQtGraph plots"""
        try:
            # Extract position from SignalProxy
            if isinstance(event, tuple):
                pos = event[0]
            else:
                pos = event

            # Determine which plot the mouse is over
            plot_item = None
            if hasattr(self, 'top_plot') and self.top_plot is not None:
                if self.top_plot.plotItem.sceneBoundingRect().contains(pos):
                    plot_item = self.top_plot.plotItem
            if hasattr(self, 'bottom_plot') and self.bottom_plot is not None:
                if self.bottom_plot.plotItem.sceneBoundingRect().contains(pos):
                    plot_item = self.bottom_plot.plotItem

            if plot_item is not None:
                # Map scene coordinates to plot coordinates
                mouse_point = plot_item.vb.mapSceneToView(pos)
                x_val = mouse_point.x()

                # ============================
                # Update cursor readout widget
                # ============================
                if hasattr(self, 'cursor_time_label'):
                    try:
                        dt = datetime.datetime.fromtimestamp(x_val)
                        self.cursor_time_label.setText(
                            f"Time: {dt.strftime('%H:%M:%S')}\n"
                        )

                        # -------- Top gauge --------
                        if self.top_timestamps is not None:
                            idx = np.argmin(np.abs(self.top_timestamps - x_val))
                            if idx < len(self.top_pressures):
                                p = self.top_pressures[idx]
                                t = self.top_temps[idx]
                                self.cursor_top_label.setText(
                                    f"Top: \nP = {p:.2f} psia\nT = {t:.2f} °F\n"
                                )

                        # -------- Bottom gauge --------
                        if self.bottom_timestamps is not None:
                            idx = np.argmin(np.abs(self.bottom_timestamps - x_val))
                            if idx < len(self.bottom_pressures):
                                p = self.bottom_pressures[idx]
                                t = self.bottom_temps[idx]
                                self.cursor_bottom_label.setText(
                                    f"Bottom: \nP = {p:.2f} psia\nT = {t:.2f} °F"
                                )

                    except Exception:
                        pass

                # Update both cursor lines
                if hasattr(self, 'cursor_vline_top') and self.cursor_vline_top is not None:
                    self.cursor_vline_top.setPos(x_val)
                if hasattr(self, 'cursor_vline_bottom') and self.cursor_vline_bottom is not None:
                    self.cursor_vline_bottom.setPos(x_val)

                # Convert timestamp to datetime for display
                try:
                    dt = datetime.datetime.fromtimestamp(x_val)
                    time_str = dt.strftime("%H:%M:%S")

                    status_lines = [f"Time: {time_str}"]

                    # Top gauge data
                    if hasattr(self, 'top_timestamps') and self.top_timestamps is not None:
                        idx = np.argmin(np.abs(self.top_timestamps - x_val))
                        if idx < len(self.top_pressures) and idx < len(self.top_temps):
                            top_p = self.top_pressures[idx]
                            top_t = self.top_temps[idx]
                            status_lines.append(f"Top: Pressure = {top_p:.2f} psia, Temp = {top_t:.2f} °F")

                    # Bottom gauge data
                    if hasattr(self, 'bottom_timestamps') and self.bottom_timestamps is not None:
                        idx = np.argmin(np.abs(self.bottom_timestamps - x_val))
                        if idx < len(self.bottom_pressures) and idx < len(self.bottom_temps):
                            bottom_p = self.bottom_pressures[idx]
                            bottom_t = self.bottom_temps[idx]
                            status_lines.append(f"Bottom: Pressure = {bottom_p:.2f} psia, Temp = {bottom_t:.2f} °F")

                    # Update status label (or print for debug)
                    # self.status_label.setText("\n".join(status_lines))
                    # print("\n".join(status_lines))  # optional debug

                except (ValueError, IndexError):
                    pass

        except Exception:
            pass

    def on_station_time_changed(self, item: QTableWidgetItem):
        row = item.row()
        col = item.column()

        # Only react to Start T. (2) or End T. (3)
        if col not in (2, 3):
            return

        if row >= len(self.station_timings):
            return

        text = item.text().strip()

        try:
            # Parse HH:MM:SS
            new_time = datetime.datetime.strptime(text, "%H:%M:%S").time()

            station = self.station_timings[row]

            # Preserve original date
            old_dt = station['start'] if col == 2 else station['end']
            new_dt = datetime.datetime.combine(old_dt.date(), new_time)

            if col == 2:
                station['start'] = new_dt
            else:
                station['end'] = new_dt

            # Safety: ensure start < end
            if station['start'] >= station['end']:
                raise ValueError("Start time must be before end time")

            # Update plot regions
            self.update_station_region(row)

            # Recompute statistics for this station
            self.recompute_station_stats(row)

        except Exception as e:
            # Revert invalid edit
            self.table_widget.blockSignals(True)
            old_dt = self.station_timings[row]['start' if col == 2 else 'end']
            item.setText(old_dt.strftime("%H:%M:%S"))
            self.table_widget.blockSignals(False)

    def update_station_region(self, index: int):
        if index >= len(self.plot_regions):
            return

        station = self.station_timings[index]
        start_ts = station['start'].timestamp()
        end_ts = station['end'].timestamp()

        region_top, region_bottom = self.plot_regions[index]

        if region_top:
            region_top.setRegion([start_ts, end_ts])

        if region_bottom:
            region_bottom.setRegion([start_ts, end_ts])

    def create_station_legend(self, colors, labels, parent_container):
        """Create a separate legend widget for stations"""
        # Create a frame for the legend
        legend_frame = QFrame()
        legend_frame.setFrameStyle(QFrame.Shape.Box.value | QFrame.Shadow.Raised.value)  # PyQt6 way
        legend_frame.setStyleSheet("background-color: white; padding: 0px; border: 0px solid #ccc;")

        legend_layout = QVBoxLayout(legend_frame)
        legend_layout.setContentsMargins(5, 5, 5, 5)

        # Add title
        title = QLabel("Stations")
        title.setStyleSheet("font-weight: bold; color: black")
        legend_layout.addWidget(title)

        # Add each station with its color
        for color, label in zip(colors, labels):
            station_item = QWidget()
            station_layout = QHBoxLayout(station_item)
            station_layout.setContentsMargins(0, 0, 0, 0)
            # station_layout.setSpacing(0)

            # Color indicator
            color_label = QLabel()
            color_label.setFixedSize(15, 15)
            color_label.setStyleSheet(
                f"background-color: rgba({color.red()}, {color.green()}, {color.blue()}, {color.alpha()}); border: 1px solid gray;")

            # Label text
            text_label = QLabel(label)
            text_label.setStyleSheet("color: black")
            print(label)

            station_layout.addWidget(color_label)
            station_layout.addWidget(text_label)
            station_layout.addStretch()

            legend_layout.addWidget(station_item)

        legend_layout.addStretch()

        # Add legend to the graph container (position it in a corner)
        parent_container.layout().addWidget(legend_frame, 0, Qt.AlignmentFlag.AlignTop)

    def populate_station_table(self):
        self.table_widget.setRowCount(len(self.station_timings))

        # Handle different data formats
        if isinstance(self.top_data, np.ndarray):
            top_data = self.top_data
        else:
            top_data = np.array(self.top_data, dtype=object) if self.top_data else None

        if isinstance(self.bottom_data, np.ndarray):
            bottom_data = self.bottom_data
        else:
            bottom_data = np.array(self.bottom_data, dtype=object) if self.bottom_data else None

        for row, station in enumerate(self.station_timings):
            # Basic station info
            self.table_widget.setItem(row, 0, QTableWidgetItem(station['station']))
            self.table_widget.setItem(row, 1, QTableWidgetItem(str(station.get('depth', 'N/A'))))
            self.table_widget.setItem(row, 2, QTableWidgetItem(station['start'].strftime("%H:%M:%S")))
            self.table_widget.setItem(row, 3, QTableWidgetItem(station['end'].strftime("%H:%M:%S")))

            # TVD data
            ahd = self.tvd_data.get('ahd_values', [])
            tvd = self.tvd_data.get('tvd_values', [])
            self.table_widget.setItem(row, 4, QTableWidgetItem(f"{ahd[row]:.2f}" if row < len(ahd) else "N/A"))
            self.table_widget.setItem(row, 5, QTableWidgetItem(f"{tvd[row]:.2f}" if row < len(tvd) else "N/A"))

            # Process gauge statistics
            for col_offset, data, gauge_name in [
                (6, top_data, 'top'),
                (12, bottom_data, 'bottom')
            ]:
                if data is None or len(data) == 0:
                    continue

                # Handle both structured arrays and object arrays
                if hasattr(data, 'dtype') and data.dtype.names is not None:
                    # Structured array
                    times = data['datetime']
                    pressures = data['pressure']
                    temps = data['temperature']
                else:
                    # Object array
                    times = data[:, 0]
                    pressures = data[:, 1].astype(float)
                    temps = data[:, 2].astype(float)

                start_dt = np.datetime64(station['start'])
                end_dt = np.datetime64(station['end'])

                mask = (times >= start_dt) & (times <= end_dt)
                if not np.any(mask):
                    stats = ["N/A"] * 6
                else:
                    p_slice = pressures[mask]
                    t_slice = temps[mask]
                    stats = [
                        np.max(p_slice), np.min(p_slice), np.median(p_slice),
                        np.max(t_slice), np.min(t_slice), np.median(t_slice)
                    ]

                for i, stat in enumerate(stats):
                    item = QTableWidgetItem(f"{stat:.2f}" if isinstance(stat, float) else str(stat))
                    self.table_widget.setItem(row, col_offset + i, item)

        # Auto-generate events after populating table
        try:
            self.auto_generate_events()
        except Exception as e:
            print(e)

        # Enable UI components
        for btn in [self.copy_button, self.generate_as2_button, self.copy_graphs_button, self.process_data_button]:
            btn.setEnabled(True)

        self.table_widget.resizeColumnsToContents()

    def show_file_upload(self):
        """Show file upload screen and reset state"""
        self.content_stack.setCurrentIndex(0)
        # Clear current state
        self.top_file_path = None
        self.bottom_file_path = None
        self.timesheet_file_path = None
        self.tvd_file_path = None
        self.top_data = None
        self.bottom_data = None
        self.station_timings = []
        self.tvd_data = {}
        self.events = []
        self.save_file_path = None

    def process_sgs_txt_file(self, file_path):
        """Process SGS text file with memory-efficient streaming"""
        try:
            # Use a more memory-efficient approach with numpy arrays
            data_points = []

            # Pre-allocate chunks to reduce reallocation
            chunk_size = 5000  # Smaller chunks for better memory control
            current_chunk = []

            with open(file_path, 'r') as f:
                # Find where data starts
                header_found = False
                data_start_line = 0
                lines_processed = 0

                # First pass: find header and count lines
                for i, line in enumerate(f):
                    if "Date" in line and "Time" in line and "Press" in line:
                        header_found = True
                        data_start_line = i + 2  # Skip header and unit line
                        break

                if not header_found:
                    MessageBoxWindow.message_simple(self, "Error", "Could not find data headers in file", "warning")
                    return None

                # Reset file pointer and skip to data start
                f.seek(0)
                for _ in range(data_start_line):
                    next(f)

                # Process data in memory-efficient chunks
                date_cache = {}  # Cache date strings to reduce memory

                for line in f:
                    try:
                        parts = line.strip().split()
                        if len(parts) < 4:
                            continue

                        # Parse date and time
                        date_str = parts[0].replace("-", "/")
                        time_str = parts[1]

                        # Use caching for date parsing to reduce object creation
                        date_key = f"{date_str}_{time_str}"

                        if date_key in date_cache:
                            date_time = date_cache[date_key]
                        else:
                            # Try different date formats
                            date_time = None
                            for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%y %H:%M:%S"):
                                try:
                                    date_time = datetime.datetime.strptime(
                                        f"{date_str} {time_str}", fmt
                                    )
                                    date_cache[date_key] = date_time
                                    break
                                except ValueError:
                                    continue

                            if date_time is None:
                                continue

                        # Parse numeric values directly without intermediate string manipulation
                        try:
                            # Direct conversion without creating intermediate strings
                            pressure = float(parts[2])
                            temperature = float(parts[3])

                            # Store as tuple (more memory efficient than list)
                            current_chunk.append((date_time, pressure, temperature))
                            lines_processed += 1

                            # Process chunk when it reaches chunk_size
                            if len(current_chunk) >= chunk_size:
                                # Extend using list comprehension for efficiency
                                data_points.extend(current_chunk)
                                # Clear chunk and free memory
                                current_chunk.clear()
                                # Force garbage collection if needed
                                if lines_processed % 50000 == 0:
                                    gc.collect()

                        except (ValueError, IndexError):
                            continue

                    except Exception:
                        continue

                # Add any remaining data
                if current_chunk:
                    data_points.extend(current_chunk)

            # Clear cache to free memory
            date_cache.clear()
            print('length of data points:', len(data_points))

            if len(data_points) == 0:
                return None

            # ALWAYS return as list of tuples with Python datetime objects
            # This ensures compatibility with timestamp() method
            print('process_sgs_txt_file complete (return data_points)')
            return data_points

        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"Failed to process data file:\n{str(e)}", "warning")
            return None

    def add_event(self):
        try:
            desc = self.event_desc_edit.text().strip()
            if not desc:
                MessageBoxWindow.message_simple(self, "Missing Description", "Please enter an event description",
                                                "warning")
                return

            base_date = self.event_date_edit.date().toPyDate()
            time_val = self.event_time_edit.time().toPyTime()
            event_datetime = datetime.datetime.combine(base_date, time_val)

            self.events.append((event_datetime, desc))

            row = self.event_table.rowCount()
            self.event_table.insertRow(row)
            self.event_table.setItem(row, 0, QTableWidgetItem(event_datetime.strftime("%Y-%m-%d %H:%M:%S")))
            self.event_table.setItem(row, 1, QTableWidgetItem(desc))

            self.event_desc_edit.clear()
            self.remove_event_btn.setEnabled(True)
        except Exception as e:
            print(e)

    def process_excel_timesheet(self, file_path):
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                df = pd.read_excel(file_path, header=None, engine='openpyxl')
            elif lower_path.endswith('.xls'):  # Now handles both .xls and .XLS
                df = pd.read_excel(file_path, header=None, engine='xlrd')
            else:
                MessageBoxWindow.message_simple(self, "Error", "Unsupported timesheet format. Must be .xlsx or .xls", "warning")
                return []
            # Find the date - handles both formats
            date_value = None
            for i in range(len(df)):
                cell_value = str(df.iloc[i, 2])
                if "Date" in cell_value:
                    # Main format: "Date of Survey  :  1/05/2025"
                    if "/" in cell_value:
                        date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', cell_value)
                        if date_match:
                            date_value = datetime.datetime.strptime(date_match.group(1), "%d/%m/%Y").date()
                    # Alternate format: "Date : 01.05.2025"
                    elif "." in cell_value:
                        date_match = re.search(r'Date :\s*(\d{2}\.\d{2}\.\d{4})', cell_value)
                        if date_match:
                            date_value = datetime.datetime.strptime(date_match.group(1), "%d.%m.%Y").date()
                    break

            if date_value is None:
                MessageBoxWindow.message_simple(self, "Error", "Could not find survey date in timesheet", "warning")
                return []

            # Determine format and find data start
            format_type = None
            data_start = None

            # Check for main format (has "No" in D and "Depth (ft)" in E)
            for i in range(len(df)):
                col_d = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""
                col_e = str(df.iloc[i, 4]).strip() if pd.notna(df.iloc[i, 4]) else ""

                if col_d == 'No' and col_e == 'Depth (ft)':
                    format_type = 'main'
                    data_start = i + 1  # Data starts next row
                    break

            # Check for alternate format (has "ATM" in D)
            if format_type is None:
                for i in range(len(df)):
                    col_d = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""
                    if col_d == 'ATM':
                        format_type = 'alternate'
                        data_start = i
                        break

            if format_type is None or data_start is None:
                MessageBoxWindow.message_simple(self, "Error", "Could not determine timesheet format", "warning")
                return []

            # Parse station data based on format
            stations = []
            i = data_start
            self.start_time = df.iloc[i, 6]
            self.time_label.setText("Start Time\t: " + str(self.start_time))

            while i < len(df):
                # Main format columns
                if format_type == 'main':
                    station_type = str(df.iloc[i, 4]).strip() if pd.notna(df.iloc[i, 4]) else ""  # Column E
                    depth = station_type  # Depth is in same column
                    duration = df.iloc[i, 5]  # Column F
                    start_time = df.iloc[i, 6]  # Column G
                    end_time = df.iloc[i, 8]  # Column I

                    # Skip ATM stations
                    if station_type == 'ATM':
                        i += 1
                        continue

                    # Set THP depth to 0
                    if station_type == 'THP':
                        depth = 0

                # Alternate format columns
                else:  # format_type == 'alternate'
                    station_type = str(df.iloc[i, 3]).strip() if pd.notna(df.iloc[i, 3]) else ""  # Column D
                    depth = df.iloc[i, 4]  # Column E
                    duration = df.iloc[i, 5]  # Column F
                    start_time = df.iloc[i, 6]  # Column G
                    end_time = df.iloc[i, 8]  # Column I

                    # Skip ATM stations
                    if station_type == 'ATM':
                        i += 1
                        continue

                    # Set THP depth to 0
                    if station_type == 'THP':
                        depth = 0

                # Skip if any critical value is missing
                if (not station_type or
                        pd.isna(duration) or
                        pd.isna(start_time) or
                        pd.isna(end_time)):
                    i += 1
                    continue

                # Convert times to time objects
                start_time_obj = self.parse_time(start_time)
                end_time_obj = self.parse_time(end_time)

                # Create datetime objects
                try:
                    # Track date rollover
                    if 'current_date' not in locals():
                        current_date = date_value
                        previous_time = start_time_obj
                    else:
                        # If the current start time is earlier than previous, it's likely past midnight
                        if start_time_obj < previous_time:
                            current_date += datetime.timedelta(days=1)
                        previous_time = start_time_obj

                    start_dt = datetime.datetime.combine(current_date, start_time_obj)
                    end_dt = datetime.datetime.combine(current_date, end_time_obj)

                    # If end time is earlier than start time, it rolled past midnight too
                    if end_dt < start_dt:
                        end_dt += datetime.timedelta(days=1)

                except Exception as e:
                    print(f"Error creating datetime: {e}")
                    i += 1
                    continue

                stations.append({
                    'station': station_type,
                    'depth': depth,
                    'start': start_dt,
                    'end': end_dt,
                    'duration': duration
                })
                i += 1

            return stations

        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"Failed to process timesheet:\n{str(e)}", "warning")
            return []



    # Optimized data processing pipeline
    def process_data(self):
        """Streamlined data processing pipeline"""
        try:
            self._get_template_path()
            self.copy_statistics()
            self.paste_to_template(self.template_path)
            self.generate_as2_files()

            MessageBoxWindow.message_simple(self, "Processing Complete",
                "Data processed successfully!\n\n"
                f"Template saved as: {os.path.basename(self.template_path)}\n"
                "AS2 files generated for both gauges")
        except Exception as e:
            MessageBoxWindow.message_simple(self, "Processing Error", f"Failed to process data:\n{str(e)}", "warning")

    def process_all_files(self, top_file_path, bottom_file_path, timesheet_file_path, tvd_file_path):
        try:
            self.cleanup_memory()
            self.top_file_path = top_file_path
            self.bottom_file_path = bottom_file_path
            self.timesheet_file_path = timesheet_file_path
            self.tvd_file_path = tvd_file_path

            # Process TVD file first
            self.tvd_data = self.process_tvd_file(tvd_file_path)

            # Process timesheet to get station timings
            self.station_timings = self.process_excel_timesheet(timesheet_file_path)

            # Process both gauge data files
            self.top_data = self.process_sgs_txt_file(top_file_path)
            self.bottom_data = self.process_sgs_txt_file(bottom_file_path)

            # Reset events when processing new files
            self.events = []
            self.event_table.setRowCount(0)
            self.remove_event_btn.setEnabled(False)
            print(1)

            # Fix: Properly check if data exists for numpy arrays and lists
            has_top_data = self.top_data is not None and (
                    (isinstance(self.top_data, np.ndarray) and len(self.top_data) > 0) or
                    (isinstance(self.top_data, list) and len(self.top_data) > 0)
            )
            print(2)

            has_bottom_data = self.bottom_data is not None and (
                    (isinstance(self.bottom_data, np.ndarray) and len(self.bottom_data) > 0) or
                    (isinstance(self.bottom_data, list) and len(self.bottom_data) > 0)
            )
            print(3)

            has_station_timings = self.station_timings and len(self.station_timings) > 0
            has_tvd_data = self.tvd_data and len(self.tvd_data.get('ahd_values', [])) > 0
            print(4)

            if has_top_data and has_bottom_data and has_station_timings and has_tvd_data:
                # Show results screen
                print(5)
                self.content_stack.setCurrentIndex(1)
                print(6)

                # Create stacked graphs
                self.show_stacked_graphs(top_file_path, bottom_file_path)
                print(7)

                # Populate table with station timings and gauge data
                self.populate_station_table()
            else:
                # Log what's missing for debugging
                missing_items = []
                if not has_top_data:
                    missing_items.append("Top Gauge Data")
                if not has_bottom_data:
                    missing_items.append("Bottom Gauge Data")
                if not has_station_timings:
                    missing_items.append("Station Timings")
                if not has_tvd_data:
                    missing_items.append("TVD Data")

                MessageBoxWindow.message_simple(
                    self,
                    "Error",
                    f"Failed to process one or more files.\nMissing: {', '.join(missing_items)}",
                    "warning"
                )

        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"Failed to process files:\n{str(e)}", "warning")

    def recompute_station_stats(self, row: int):
        """
        Recompute high / low / median pressure & temperature
        for a single station and update the table row.
        """

        if row >= len(self.station_timings):
            return

        station = self.station_timings[row]
        start_dt = station['start']
        end_dt = station['end']

        start_np = np.datetime64(start_dt)
        end_np = np.datetime64(end_dt)

        def compute_stats(data):
            if data is None or len(data) == 0:
                return ["N/A"] * 6

            # Structured array
            if hasattr(data, 'dtype') and data.dtype.names is not None:
                times = data['datetime']
                pressures = data['pressure']
                temps = data['temperature']
            else:
                # List-of-tuples format
                data_np = np.array(data, dtype=object)
                times = data_np[:, 0]
                pressures = data_np[:, 1].astype(float)
                temps = data_np[:, 2].astype(float)

            mask = (times >= start_np) & (times <= end_np)

            if not np.any(mask):
                return ["N/A"] * 6

            p_slice = pressures[mask]
            t_slice = temps[mask]

            return [
                np.max(p_slice), np.min(p_slice), np.median(p_slice),
                np.max(t_slice), np.min(t_slice), np.median(t_slice)
            ]

        # Compute new stats
        top_stats = compute_stats(self.top_data)
        bottom_stats = compute_stats(self.bottom_data)

        self.table_widget.blockSignals(True)
        # Write back to table
        for i, stat in enumerate(top_stats):
            col = 6 + i
            self.table_widget.setItem(
                row, col,
                QTableWidgetItem(f"{stat:.2f}" if isinstance(stat, float) else str(stat))
            )

        for i, stat in enumerate(bottom_stats):
            col = 12 + i
            self.table_widget.setItem(
                row, col,
                QTableWidgetItem(f"{stat:.2f}" if isinstance(stat, float) else str(stat))
            )
        self.table_widget.blockSignals(False)

    def get_cell_value(self, sheet, row, col):
        """Unified cell value getter for different Excel formats"""
        try:
            if isinstance(sheet, openpyxl.worksheet.worksheet.Worksheet):
                return sheet.cell(row=row+1, column=col+1).value
            elif isinstance(sheet, xlrd.sheet.Sheet):
                return sheet.cell_value(row, col)
        except (IndexError, AttributeError):
            return None

    def parse_numeric_cell(self, value):
        """Parse numeric cell value with error handling"""
        if value is None:
            return None
        try:
            if isinstance(value, str):
                value = value.replace(',', '')
            return float(value)
        except (ValueError, TypeError):
            return None


    def parse_time(self, time_val):
        """Parse time values from various formats into time objects"""
        if isinstance(time_val, datetime.time):
            return time_val

        if isinstance(time_val, str):
            time_val = time_val.strip()
            # Handle colon-separated times (12:17:00)
            if ':' in time_val:
                parts = time_val.split(':')
                hour = int(parts[0])
                minute = int(parts[1])
                second = int(parts[2]) if len(parts) >= 3 else 0
                return datetime.time(hour, minute, second)
            # Handle 4-digit times (1217)
            else:
                # Remove non-digits and pad with zeros
                digits = ''.join(filter(str.isdigit, time_val)).zfill(4)[:4]
                hour = int(digits[:2])
                minute = int(digits[2:4])
                return datetime.time(hour, minute)

        elif isinstance(time_val, (int, float)):
            # Handle integer times (1217)
            time_val = int(time_val)
            hour = time_val // 100
            minute = time_val % 100
            return datetime.time(hour, minute)

        # Return midnight as default
        return datetime.time(0, 0)

    def process_tvd_file(self, file_path):
        """Optimized TVD file processing with better validation"""
        try:
            lower_path = file_path.lower()
            if lower_path.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheet = wb[wb.sheetnames[0]]
            elif lower_path.endswith('.xls'):
                wb = xlrd.open_workbook(file_path)
                sheet = wb.sheet_by_index(0)
            else:
                MessageBoxWindow.message_simple(self, "Error", "Unsupported TVD file format", "warning")
                return {}

            tvd_data = {
                'ahd_values': [],
                'tvd_values': []
            }
            self.bdf = self.get_cell_value(sheet, 1, 3)
            self.sea_level = self.get_cell_value(sheet, 2, 3)
            self.date = self.get_cell_value(sheet, 1, 8)
            self.location = self.get_cell_value(sheet, 2, 8)
            self.well = self.get_cell_value(sheet, 3, 8)

            self.location_label.setText("Location\t: " + str(self.location))
            self.well_label.setText("Well No.\t: " + str(self.well))
            self.bdf_label.setText("THF\t\t: " + str(self.bdf) + " ft BDF")
            self.sea_level_label.setText("DFE\t\t: " + str(self.sea_level) + " ft AMSL")
            self.date_label.setText("Date of Survey\t: " + self.date.strftime("%d/%m/%Y"))

            if hasattr(self, 'event_date_edit'):  # If UI is already created
                self.event_date_edit.setDate(QDate(self.date))

            for row in range(8, sheet.max_row if hasattr(sheet, 'max_row') else sheet.nrows):
                ahd_val = self.parse_numeric_cell(self.get_cell_value(sheet, row, 2))
                tvd_val = self.parse_numeric_cell(self.get_cell_value(sheet, row, 8))

                if ahd_val is None and tvd_val is None:
                    break

                if ahd_val is not None:
                    tvd_data['ahd_values'].append(ahd_val)
                if tvd_val is not None:
                    tvd_data['tvd_values'].append(tvd_val)

            return tvd_data

        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"TVD Processing Error:\n{str(e)}", "warning")
            return {}


    def reset_button(self, button, original_text, original_icon):
        """Revert button to original state"""
        # Determine the appropriate style based on button type
        if button == self.copy_button:
            style = """
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    border-radius: 5px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #2980b9;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #888888;
                }
            """
        elif button == self.copy_graphs_button:
            style = """
                QPushButton {
                    background-color: #9b59b6;
                    color: white;
                    border-radius: 5px;
                    padding: 8px;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #8e44ad;
                }
                QPushButton:disabled {
                    background-color: #cccccc;
                    color: #888888;
                }
            """
        else:
            style = ""

        button.setStyleSheet(style)
        button.setIcon(QIcon(get_icon_path(original_icon)))
        button.setText(original_text)

    def show_file_upload(self):
        # Switch back to file upload screen
        self.content_stack.setCurrentIndex(0)

    def get_ordinal(self, n):
        """Convert integer to ordinal string (1st, 2nd, 3rd, etc.)"""
        if 10 <= n % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
        return f"{n}{suffix}"

    def copy_graphs(self):
        """Copy the current graphs to clipboard as an image"""
        try:
            if hasattr(self, 'top_plot') and self.top_plot is not None:
                # Grab the graph widget as QPixmap
                pixmap = self.top_plot.grab()

                # Copy to clipboard
                clipboard = QApplication.clipboard()
                clipboard.setPixmap(pixmap)

                # Show feedback
                self.set_button_success_feedback(
                    self.copy_graphs_button,
                    "Copied!",
                    'check',
                    "Copy Graphs",
                    'copy'
                )
            else:
                MessageBoxWindow.message_simple(self, "No Graphs", "No graphs available to copy", "warning")
        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"Failed to copy graphs:\n{str(e)}", "warning")

    def reset_graph_zoom(self):
        """Reset graph zoom to show all data"""
        if hasattr(self, 'top_plot') and self.top_plot is not None:
            self.top_plot.autoRange()
            self.reset_zoom_btn.setEnabled(False)

    def cleanup_memory(self):
        """Clean up memory-intensive objects"""
        self.top_times = None
        self.top_pressures = None
        self.top_temps = None
        self.bottom_times = None
        self.bottom_pressures = None
        self.bottom_temps = None

        if hasattr(self, 'current_canvas') and self.current_canvas:
            self.current_canvas.figure.clear()
            # plt.close(self.current_canvas.figure)
            self.current_canvas = None

        gc.collect()

    def toggle_theme(self):
        self.current_theme = toggle_theme(
            widget=self,
            current_theme=self.current_theme,
            theme_button=self.theme_button,
            summary_widget=None
        )

    def remove_event(self):
        """Remove selected event from the events list and table"""
        selected_row = self.event_table.currentRow()
        if selected_row >= 0:
            self.event_table.removeRow(selected_row)
            del self.events[selected_row]

            # Disable remove button if no events left
            if not self.events:
                self.remove_event_btn.setEnabled(False)

    def _download_template(self, template_name, dialog_title, silent=False):
        """Generic template download handler with TVD directory as default"""
        try:
            template_path = get_path(f"assets/resources/{template_name}")

            if silent:
                temp_path = os.path.join(tempfile.gettempdir(), os.path.basename(template_name))
                shutil.copy(template_path, temp_path)
                return temp_path

            # Set default directory to TVD file location if available
            default_dir = ""
            if hasattr(self, 'tvd_file_path') and self.tvd_file_path:
                default_dir = os.path.dirname(self.tvd_file_path)
            else:
                default_dir = os.getcwd()

            file_path, _ = QFileDialog.getSaveFileName(
                self,
                dialog_title,
                os.path.join(default_dir, os.path.basename(template_name)),
                "Excel Files (*.xlsx)"
            )

            if file_path:
                shutil.copy(template_path, file_path)
                if not silent:
                    MessageBoxWindow.message_simple(self, "Template Downloaded", f"Template saved to:\n{file_path}")
                return file_path
            return None

        except Exception as e:
            if not silent:
                MessageBoxWindow.message_simple(self, "Error", f"Failed to download template:\n{str(e)}")
            return None

    def download_md_tvd_template(self):
        """Download the MD-to-TVD template Excel file"""
        return self._download_template(
            "MD_TVD_Template.xlsx",
            "Save MD-to-TVD Template"
        )


    def download_timesheet_template(self):
        """Download the MD-to-TVD template Excel file"""
        return self._download_template(
            "Timesheet_Template.xlsx",
            "Save Timesheet Template"
        )

    def download_interpretation_template(self, silent=False):
        """Download interpretation template, optionally return path without dialog"""
        return self._download_template(
            "Interpretation_Template.xlsx",
            "Save Interpretation Template",
            silent
        )

    def _get_template_path(self):
        """Helper for template path handling"""
        self.template_path = self.download_interpretation_template(silent=False)
        return bool(self.template_path)

    def generate_as2_files(self):
        """Generate AS2 files for both top and bottom gauges"""
        if not self.top_data or not self.bottom_data:
            MessageBoxWindow.message_simple(self, "No Data", "No gauge data available to generate AS2 files", "warning")
            return

        try:
            top_as2_path = self.generate_as2_file(self.top_file_path, self.top_data)
            bottom_as2_path = self.generate_as2_file(self.bottom_file_path, self.bottom_data)

            if top_as2_path and bottom_as2_path:
                MessageBoxWindow.message_simple(
                    self,
                    "Files Created",
                    f"Successfully created AS2 files:\n\n"
                    f"Top Gauge: {top_as2_path.split('/')[-1]}\n"
                    f"Bottom Gauge: {bottom_as2_path.split('/')[-1]}",
                    "check_green")
        except Exception as e:
            MessageBoxWindow.message_simple(self, "Error", f"Failed to create AS2 files:\n{str(e)}", "warning")

    def generate_as2_file(self, input_file_path, data):
        """Generate AS2 formatted file from processed data with right-aligned columns"""
        # Determine output file path
        if input_file_path.endswith('.txt'):
            output_file_path = input_file_path.replace('.txt', '.AS2')
        else:
            output_file_path = input_file_path + '.AS2'

        if not data:
            # Create empty file if no data
            with open(output_file_path, 'w'):
                pass
            return output_file_path

        # Create event lookup dictionary
        event_dict = {dt.strftime("%Y-%m-%d %H:%M:%S"): desc for dt, desc in self.events}


        # Calculate column widths
        max_pressure_width = max(len(f"{p:.2f}") for _, p, _ in data) if data else 0
        max_temp_width = max(len(f"{t:.3f}") for _, _, t in data) if data else 0

        with open(output_file_path, 'w') as f:
            for dt, pressure, temperature in data:
                date_str = dt.strftime("%d/%m/%Y")
                time_str = dt.strftime("%H:%M:%S")
                datetime_str = dt.strftime("%Y-%m-%d %H:%M:%S")

                temp_str = f"{temperature:.2f}".rjust(max_temp_width)
                pressure_str = f"{pressure:.3f}".rjust(max_pressure_width)

                event_desc = event_dict.get(datetime_str, "")
                line = f"{date_str}  {time_str}    {temp_str}     {pressure_str}  {event_desc}\n"
                f.write(line)

        return output_file_path

    def save_file(self):
        """Save current application state to a file"""
        if not self.station_timings:
            MessageBoxWindow.message_simple(self, "No Data", "Nothing to save - process files first", "warning")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Survey", "", "Survey Files (*.survey)"
        )
        if not file_path:
            return

        # Add extension if needed
        if not file_path.endswith('.survey'):
            file_path += '.survey'

        try:
            # Prepare state dictionary
            state = {
                'top_file_path': self.top_file_path,
                'bottom_file_path': self.bottom_file_path,
                'timesheet_file_path': self.timesheet_file_path,
                'tvd_file_path': self.tvd_file_path,
                'top_data': self.top_data,
                'bottom_data': self.bottom_data,
                'station_timings': self.station_timings,
                'tvd_data': self.tvd_data,
                'events': self.events,
                'location': self.location,
                'well': self.well,
                'date': self.date,
                'start_time': self.start_time,
                'bdf': self.bdf,
                'sea_level': self.sea_level,
                'app_version': '1.0'
            }

            with open(file_path, 'wb') as f:
                pickle.dump(state, f)

            self.save_file_path = file_path
            MessageBoxWindow.message_simple(
                self, "Save Successful",
                f"Survey saved successfully to:\n{file_path}"
            )
        except Exception as e:
            MessageBoxWindow.message_simple(
                self, "Save Error",
                f"Failed to save survey:\n{str(e)}",
                "warning"
            )

    def open_file(self):
        """Load application state from a file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Survey", "", "Survey Files (*.survey)"
        )
        if not file_path:
            return

        try:
            with open(file_path, 'rb') as f:
                state = pickle.load(f)

            # Validate file version
            if state.get('app_version') != '1.0':
                MessageBoxWindow.message_simple(
                    self, "Version Mismatch",
                    "This file was created with a different version of the application",
                    "warning"
                )
                return

            # Load state into application
            self.top_file_path = state.get('top_file_path')
            self.bottom_file_path = state.get('bottom_file_path')
            self.timesheet_file_path = state.get('timesheet_file_path')
            self.tvd_file_path = state.get('tvd_file_path')
            self.top_data = state.get('top_data')
            self.bottom_data = state.get('bottom_data')
            self.station_timings = state.get('station_timings')
            self.tvd_data = state.get('tvd_data')
            self.events = state.get('events', [])
            self.location = state.get('location')
            self.well = state.get('well')
            self.date = state.get('date')
            self.start_time = state.get('start_time')
            self.bdf = state.get('bdf')
            self.sea_level = state.get('sea_level')
            self.save_file_path = file_path

            # Update UI
            self.content_stack.setCurrentIndex(1)  # Show results screen
            self.update_info_labels()
            self.populate_station_table()
            self.populate_events_table()
            self.show_stacked_graphs(self.top_file_path, self.bottom_file_path)

            # Enable UI components
            for btn in [self.copy_button, self.generate_as2_button,
                        self.copy_graphs_button, self.process_data_button]:
                btn.setEnabled(True)

            self.reset_zoom_btn.setEnabled(True)
            self.copy_graphs_button.setEnabled(True)

            MessageBoxWindow.message_simple(
                self, "Load Successful",
                f"Survey loaded successfully from:\n{file_path}"
            )
        except Exception as e:
            MessageBoxWindow.message_simple(
                self, "Load Error",
                f"Failed to load survey:\n{str(e)}",
                "warning"
            )

    def auto_generate_events(self):
        """Auto-generate events based on processed data and station timings"""
        self.events = []  # Clear existing events

        # 1. Battery connected (earliest time from .txt file)
        if self.top_data or self.bottom_data:
            all_times = []
            if self.top_data:
                all_times.extend([d[0] for d in self.top_data])
            if self.bottom_data:
                all_times.extend([d[0] for d in self.bottom_data])
            if all_times:
                min_time = min(all_times)
                self.events.append((min_time, "Battery Connected"))

        # 2. ATM Reading (start time of ATM station)
        atm_stations = [s for s in self.station_timings if s['station'] == 'ATM']
        if atm_stations:
            first_atm = atm_stations[0]
            self.events.append((first_atm['start'], "Reading ATM"))

        # 3. Open Swab Valve (pressure increase from data)
        if self.top_data:
            # Find first significant pressure increase
            base_pressure = None
            for data_point in self.top_data:
                _, pressure, _ = data_point
                if base_pressure is None:
                    base_pressure = pressure
                elif pressure > base_pressure + 0.5:  # 500 psi threshold
                    self.events.append((data_point[0], "Open Swab Valve"))
                    break

        # 4. THP Reading and POOH events
        thp_stations = [s for s in self.station_timings if s['station'] == 'THP']
        non_thp_stations = [s for s in self.station_timings if s['station'] not in ['ATM', 'THP']]

        for idx, thp in enumerate(thp_stations):
            # THP Reading event
            self.events.append((thp['start'], "Reading THP"))

            # Only add POOH if it's not the last THP station
            first = True
            if idx < len(thp_stations) - 1:
                if first:
                    self.events.append((thp['end'], "RIH"))
                    first = False
                else:
                    self.events.append((thp['end'], "POOH"))

            # For each non-THP station after this THP
            station_count = 0
            for station in non_thp_stations:
                if station['start'] > thp['end']:
                    station_count += 1
                    ordinal = self.get_ordinal(station_count)

                    # Station start event
                    self.events.append((
                        station['start'],
                        f"{ordinal} Stop at {station['depth']} ft WLD"
                    ))

                    # Station end event (POOH)
                    self.events.append((station['end'], "POOH"))

        # Update event table UI
        self.event_table.setRowCount(0)
        for event_time, desc in self.events:
            row = self.event_table.rowCount()
            self.event_table.insertRow(row)
            self.event_table.setItem(row, 0, QTableWidgetItem(event_time.strftime("%Y-%m-%d %H:%M:%S")))
            self.event_table.setItem(row, 1, QTableWidgetItem(desc))

        # Enable remove button
        self.remove_event_btn.setEnabled(len(self.events) > 0)

    def copy_statistics(self):
        """Copy pressure and temperature statistics to clipboard with visual feedback"""
        clipboard = QApplication.clipboard()

        # Get statistics data
        stats_data = []
        for row in range(self.table_widget.rowCount()):
            row_data = []
            for col in range(4, 18):  # Columns 5 to 18
                item = self.table_widget.item(row, col)
                row_data.append(item.text() if item else "")
            stats_data.append("\t".join(row_data))

        # Format as tab-separated values
        text_data = "\n".join(stats_data)

        # Set to clipboard
        clipboard.setText(text_data)

        # Show visual feedback
        self.set_button_success_feedback(
            self.copy_button,
            "Copied!",
            'check',
            "Copy",
            'copy'
        )

    def paste_to_template(self, template_path):
        """Paste clipboard data to template starting at C13 with row deletion"""
        try:
            wb = openpyxl.load_workbook(template_path)
            sheet = wb.active
            clipboard = QApplication.clipboard()
            rows = clipboard.text().split('\n')

            # Count valid data rows (non-empty)
            x = sum(1 for row in rows if row.strip())

            # Calculate unused rows
            U = 69 - x

            # Delete unused rows if needed
            if U > 0:
                # Delete from top section (row 80-U to 80)
                start_top = 82 - U
                sheet.delete_rows(start_top, U + 1)  # +1 to include end row

            # Paste data
            for row_idx, row in enumerate(rows):
                if not row.strip():
                    continue

                cells = row.split('\t')[:14]  # Only process first 14 columns
                for col_idx, cell_value in enumerate(cells):
                    try:
                        value = float(cell_value)
                    except ValueError:
                        value = cell_value

                    sheet.cell(
                        row=13 + row_idx,
                        column=3 + col_idx,
                        value=value
                    )

            # Update header information
            sheet.cell(row=3, column=3, value=f": {self.location}")
            sheet.cell(row=4, column=3, value=f": {self.well}")
            sheet.cell(row=3, column=18, value=self.date)
            sheet.cell(row=4, column=18, value=self.sea_level)
            sheet.cell(row=5, column=18, value=self.bdf)

            wb.save(template_path)

            # Access the first chart
            chart = sheet._charts[0]

            # --- Get Pressure Min/Max (Column G) ---
            pressure_values = [
                sheet.cell(row=r, column=7).value
                for r in range(13, 13 + x)
                if isinstance(sheet.cell(row=r, column=7).value, (int, float))
            ]
            if pressure_values:
                chart.x_axis.scaling.min = min(pressure_values)
                chart.x_axis.scaling.max = max(pressure_values)

            # --- Get Temperature Min/Max (Column E) ---
            temp_values = [
                sheet.cell(row=r, column=10).value
                for r in range(13, 13 + x)
                if isinstance(sheet.cell(row=r, column=10).value, (int, float))
            ]
            # if temp_values:
            #     if chart.x2_axis.secondary_axis:
            #         chart.x2_axis.scaling.min = min(temp_values)
            #         chart.x2_axis.secondary_axis.scaling.max = max(temp_values)

            # Save the workbook again after modifying chart axes
            wb.save(template_path)

            return True

        except Exception as e:
            MessageBoxWindow.message_simple(self, "Paste Error", f"Failed to paste to template:\n{str(e)}", "warning")
            return False

    # Add this helper method to the class
    def set_button_success_feedback(self, button, success_text, success_icon, original_text, original_icon):
        """Set button to success state and schedule reset"""
        button.setStyleSheet(f"""
            QPushButton {{
                background-color: #28a745;
                color: white;
                border-radius: 5px;
                padding: 8px;
                font-weight: bold;
            }}
        """)
        button.setIcon(QIcon(get_icon_path(success_icon)))
        button.setText(success_text)

        # Set timer to revert button after 3 seconds
        QTimer.singleShot(3000, lambda: self.reset_button(button, original_text, original_icon))
